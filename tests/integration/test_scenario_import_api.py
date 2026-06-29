from io import BytesIO

import pytest
from openpyxl import Workbook

from app.services.scenario_import import ParsedRow, _parse_effects


def test_scenario_import_effect_parser_accepts_sub_and_reset():
    row = ParsedRow(sheet="rules", row_number=2, data={})
    errors = []

    effects = _parse_effects("cleanliness:sub:25;cleanliness:reset:100", row, errors)

    assert errors == []
    assert effects == [
        {
            "feature_key": "cleanliness",
            "effect_type": "sub",
            "new_value": "25",
            "delta_value": 25.0,
        },
        {
            "feature_key": "cleanliness",
            "effect_type": "reset",
            "new_value": "100",
            "delta_value": None,
        },
    ]


def _scenario_workbook(*, broken_effect: bool = False) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {
        "meta": ["scenario_code", "scenario_name", "version", "mode"],
        "feature_catalog": ["feature_key", "value_type", "allowed_values", "unit", "description"],
        "machine_type": ["code", "name", "description"],
        "machines": ["code", "machine_type_code", "name", "location"],
        "state_feature_defs": ["machine_type_code", "feature_key", "feature_name", "value_type", "allowed_values"],
        "resources": ["machine_code", "code", "name", "resource_type", "capacity", "is_available", "meta_json"],
        "rules": [
            "code",
            "machine_type_code",
            "name",
            "duration_min",
            "description",
            "is_active",
            "is_repair",
            "preconditions",
            "effects",
            "resource_reqs",
        ],
        "states": ["machine_code", "state_code", "state_type", "label", "features"],
        "solve_cases": [
            "case_code",
            "machine_code",
            "current_state_code",
            "target_state_code",
            "objective",
            "objectives_json",
            "constraints_json",
            "expected_min_steps",
            "expected_max_makespan_min",
        ],
        "instructions": ["section", "description"],
    }
    for name, header in sheets.items():
        sheet = workbook.create_sheet(name)
        sheet.append(header)

    workbook["meta"].append(["E2E_SCENARIO", "Scenario Import E2E", "v1", "scenario_upsert"])
    workbook["feature_catalog"].append(["prep_done", "enum", "false,true", "", ""])
    workbook["feature_catalog"].append(["delivery_ready", "enum", "false,true", "", ""])
    workbook["machine_type"].append(["IMPORT_MACHINE", "Import Machine", ""])
    workbook["machines"].append(["IMPORT-001", "IMPORT_MACHINE", "Import Test Machine", "Line A"])
    workbook["state_feature_defs"].append(["IMPORT_MACHINE", "prep_done", "Prep Done", "enum", "false,true"])
    workbook["state_feature_defs"].append(["IMPORT_MACHINE", "delivery_ready", "Delivery Ready", "enum", "false,true"])
    workbook["resources"].append(["IMPORT-001", "TECH-IMPORT-01", "Import Tech 01", "TECHNICIAN", 1, "true", ""])
    workbook["resources"].append(["IMPORT-001", "QA-IMPORT-01", "Import QA 01", "QA", 1, "true", ""])
    workbook["rules"].append([
        "OP_IMPORT_PREP",
        "IMPORT_MACHINE",
        "Prepare",
        10,
        "",
        "true",
        "false",
        "prep_done:eq:false",
        "prep_done:set:true",
        "TECHNICIAN:1:true",
    ])
    workbook["rules"].append([
        "OP_IMPORT_DELIVER",
        "IMPORT_MACHINE",
        "Deliver",
        15,
        "",
        "true",
        "false",
        "prep_done:eq:true",
        "missing_feature:set:true" if broken_effect else "delivery_ready:set:true",
        "QA:1:true",
    ])
    workbook["states"].append(["IMPORT-001", "START", "current", "Start", "prep_done:false;delivery_ready:false"])
    workbook["states"].append(["IMPORT-001", "TARGET", "target", "Target", "prep_done:true;delivery_ready:true"])
    workbook["solve_cases"].append(["FULL_FLOW", "IMPORT-001", "START", "TARGET", "minimize_makespan", "", "", 2, 25])
    workbook["instructions"].append(["example", "Fill sheets from left to right."])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _large_scenario_workbook(rule_count: int = 105) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {
        "meta": ["scenario_code", "scenario_name", "version", "mode"],
        "feature_catalog": ["feature_key", "value_type", "allowed_values", "unit", "description"],
        "machine_type": ["code", "name", "description"],
        "machines": ["code", "machine_type_code", "name", "location"],
        "state_feature_defs": ["machine_type_code", "feature_key", "feature_name", "value_type", "allowed_values"],
        "resources": ["machine_code", "code", "name", "resource_type", "capacity", "is_available", "meta_json"],
        "rules": [
            "code",
            "machine_type_code",
            "name",
            "duration_min",
            "description",
            "is_active",
            "is_repair",
            "preconditions",
            "effects",
            "resource_reqs",
        ],
        "states": ["machine_code", "state_code", "state_type", "label", "features"],
        "solve_cases": [
            "case_code",
            "machine_code",
            "current_state_code",
            "target_state_code",
            "objective",
            "objectives_json",
            "constraints_json",
            "expected_min_steps",
            "expected_max_makespan_min",
        ],
        "instructions": ["section", "description"],
    }
    for name, header in sheets.items():
        sheet = workbook.create_sheet(name)
        sheet.append(header)

    workbook["meta"].append(["LARGE_SCENARIO", "Large Scenario", "v1", "scenario_upsert"])
    workbook["machine_type"].append(["LARGE_MACHINE", "Large Machine", ""])
    workbook["machines"].append(["LARGE-001", "LARGE_MACHINE", "Large Test Machine", "Line L"])
    workbook["resources"].append(["LARGE-001", "TECH-LARGE-01", "Large Tech 01", "TECHNICIAN", 1, "true", ""])

    start_features = []
    target_features = []
    for index in range(rule_count):
        feature_key = f"step_{index:03d}_done"
        workbook["feature_catalog"].append([feature_key, "enum", "false,true", "", ""])
        workbook["state_feature_defs"].append(["LARGE_MACHINE", feature_key, feature_key, "enum", "false,true"])
        precondition = "" if index == 0 else f"step_{index - 1:03d}_done:eq:true"
        workbook["rules"].append([
            f"OP_LARGE_{index:03d}",
            "LARGE_MACHINE",
            f"Large Step {index:03d}",
            5,
            "",
            "true",
            "false",
            precondition,
            f"{feature_key}:set:true",
            "TECHNICIAN:1:true",
        ])
        start_features.append(f"{feature_key}:false")
        target_features.append(f"{feature_key}:true")

    workbook["states"].append(["LARGE-001", "START", "current", "Start", ";".join(start_features)])
    workbook["states"].append(["LARGE-001", "TARGET", "target", "Target", ";".join(target_features)])
    workbook["solve_cases"].append(["LARGE_FLOW", "LARGE-001", "START", "TARGET", "minimize_makespan", "", "", rule_count, rule_count * 5])
    workbook["instructions"].append(["example", "Large scenario dry-run fixture."])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.mark.asyncio
async def test_scenario_import_dry_run_import_and_solve(client):
    content = _scenario_workbook()
    files = {
        "file": (
            "scenario.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    data = {"mode": "scenario_upsert", "dry_run": "true"}
    resp = await client.post("/api/v1/imports/scenario", data=data, files=files)
    assert resp.status_code == 200
    dry_run = resp.json()
    assert dry_run["status"] == "validated"
    assert dry_run["summary"]["rules_total"] == 2
    assert dry_run["summary"]["error_count"] == 0
    assert dry_run["preview"]["rules"] == {"create": 2, "update": 0}

    files = {
        "file": (
            "scenario.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    resp = await client.post("/api/v1/imports/scenario", data={"mode": "scenario_upsert", "dry_run": "false"}, files=files)
    assert resp.status_code == 200
    imported = resp.json()
    assert imported["status"] == "imported"
    solve_case = imported["solve_cases"][0]
    assert solve_case["machine_id"]
    assert solve_case["current_state_id"]
    assert solve_case["target_state_id"]

    solve_resp = await client.post("/api/v1/solve", json={
        "machine_id": solve_case["machine_id"],
        "current_state_id": solve_case["current_state_id"],
        "target_state_id": solve_case["target_state_id"],
        "objective": "minimize_makespan",
    })
    assert solve_resp.status_code == 200
    solve_data = solve_resp.json()
    assert solve_data["status"] == "done"
    assert solve_data["schedule"]["makespan"] == 25
    assert len(solve_data["schedule"]["tasks"]) == 2


@pytest.mark.asyncio
async def test_scenario_import_validation_error_does_not_write(client):
    files = {
        "file": (
            "broken.xlsx",
            _scenario_workbook(broken_effect=True),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    resp = await client.post("/api/v1/imports/scenario", data={"mode": "scenario_upsert", "dry_run": "false"}, files=files)
    assert resp.status_code == 200
    payload = resp.json()
    assert payload["status"] == "failed"
    assert payload["summary"]["error_count"] >= 1
    assert any(error["field"] == "effects" for error in payload["errors"])

    machines_resp = await client.get("/api/v1/machines")
    assert machines_resp.status_code == 200
    assert machines_resp.json() == []


@pytest.mark.asyncio
async def test_scenario_import_dry_run_accepts_100_plus_rules(client):
    files = {
        "file": (
            "large.xlsx",
            _large_scenario_workbook(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    resp = await client.post("/api/v1/imports/scenario", data={"mode": "scenario_upsert", "dry_run": "true"}, files=files)
    assert resp.status_code == 200
    payload = resp.json()
    assert payload["status"] == "validated"
    assert payload["summary"]["rules_total"] == 105
    assert payload["summary"]["error_count"] == 0
    assert payload["preview"]["rules"] == {"create": 105, "update": 0}


@pytest.mark.asyncio
async def test_scenario_template_download(client):
    resp = await client.get("/api/v1/imports/scenario-template")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert len(resp.content) > 1000
