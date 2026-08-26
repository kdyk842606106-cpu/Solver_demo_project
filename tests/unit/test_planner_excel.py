from io import BytesIO

from openpyxl import load_workbook

from app.services.planner_excel import export_workbook, import_workbook, template_workbook
from app.services.planner_scenarios import (
    add_membership,
    create_activity,
    create_package,
    new_scenario,
    rebuild_mirror,
    validate_scenario,
)


def _scenario():
    scenario = new_scenario("Excel 往返")
    root = create_package(scenario, {"name": "一级"}, display_number=1)
    child = create_package(scenario, {"name": "二级", "parent_id": root["id"]}, display_number=2)
    seed_id = "state:excel:seed"
    scenario["states"].append({"id": seed_id, "name": "开始", "state_kind": "seed"})
    scenario["initial_state_ids"] = [seed_id]
    activity = create_activity(
        scenario,
        {"name": "Excel 活动", "duration": 4, "preconditions": [{"state_id": seed_id, "relation_role": "transition"}]},
        display_number=1,
    )
    add_membership(scenario, child["id"], activity["id"])
    scenario["target_activity_package_ids"] = [root["id"]]
    rebuild_mirror(scenario)
    return scenario


def test_excel_template_and_round_trip_use_temporary_refs_not_solver_ids():
    template = template_workbook()
    assert len(template) > 5000
    workbook = load_workbook(BytesIO(template), data_only=True)
    assert "里程碑" not in [cell.value for cell in workbook["活动"][1]]
    assert "是否目标活动" not in [cell.value for cell in workbook["活动"][1]]
    imported = import_workbook(export_workbook(_scenario()))
    assert validate_scenario(imported) == []
    assert imported["activities"][0]["id"].startswith("activity:")
    assert imported["activity_packages"][0]["id"].startswith("activity-package:")
    assert imported["state_packages"][0]["managed_by"] == "activity_package_mirror"
    assert imported["target_activity_package_ids"] == []
    assert imported["target_activity_ids"] == []
    assert imported["activities"][0]["is_milestone"] is False


def test_legacy_excel_target_rows_are_converted_to_goal_states():
    workbook = load_workbook(BytesIO(template_workbook()))
    workbook["活动"].cell(row=1, column=5, value="是否目标活动")
    workbook["活动"].cell(row=2, column=5, value="是")
    workbook["包目标范围"].append(["target", "P001"])
    stream = BytesIO()
    workbook.save(stream)

    imported = import_workbook(stream.getvalue())

    assert imported["target_activity_ids"] == []
    assert imported["target_activity_package_ids"] == []
    assert imported["goal_state_ids"] == [imported["activities"][0]["output_state_id"]]
