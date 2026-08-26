"""Business-facing Excel round trip for Planner shared scenarios."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.services.planner_scenarios import PlannerScenarioError, normalize_import


SHEETS = {
    "场景": ["场景名称", "执行模式", "起始时间", "最大活动实例数", "时间预算秒", "转换预算"],
    "活动包": ["包引用", "名称", "父包引用", "排序"],
    "活动": ["活动引用", "名称", "工期", "最大实例数", "是否目标活动"],
    "包成员": ["包引用", "活动引用", "排序"],
    "种子状态": ["状态引用", "名称", "初始激活", "目标必需", "目标禁止"],
    "活动前置": ["活动引用", "状态引用", "关系角色"],
    "资源": ["资源引用", "名称", "容量"],
    "活动资源": ["活动引用", "资源引用", "需求数量"],
    "外部事件": ["事件引用", "名称", "发生时间"],
    "事件状态": ["事件引用", "操作", "状态引用"],
    "活动事件": ["活动引用", "事件引用"],
    "包目标范围": ["用途", "包引用"],
}


def template_workbook() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, headers in SHEETS.items():
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        _style(sheet, len(headers))
    workbook["场景"].append(["示例场景", "serial", 0, 20, 5, 20000])
    workbook["活动包"].append(["P001", "一级包", "", 10])
    workbook["活动包"].append(["P002", "二级包", "P001", 10])
    workbook["种子状态"].append(["S001", "开始", "是", "否", "否"])
    workbook["活动"].append(["A001", "示例活动", 10, "", "否"])
    workbook["包成员"].append(["P002", "A001", 10])
    workbook["活动前置"].append(["A001", "S001", "transition"])
    return _bytes(workbook)


def export_workbook(scenario: dict[str, Any]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, headers in SHEETS.items():
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        _style(sheet, len(headers))
    workbook["场景"].append([
        scenario["name"], scenario.get("execution_mode", "serial"), scenario.get("start_time", 0),
        scenario.get("max_steps", 20), scenario.get("default_budget", {}).get("time_limit_seconds", 5),
        scenario.get("default_budget", {}).get("transition_limit", 20000),
    ])
    package_refs = {item["id"]: f"P{index:03d}" for index, item in enumerate(scenario.get("activity_packages", []), 1)}
    activity_refs = {item["id"]: f"A{index:03d}" for index, item in enumerate(scenario.get("activities", []), 1)}
    resource_refs = {item["id"]: f"R{index:03d}" for index, item in enumerate(scenario.get("resources", []), 1)}
    event_refs = {item["id"]: f"E{index:03d}" for index, item in enumerate(scenario.get("external_events", []), 1)}
    state_refs: dict[str, str] = {}
    seed_states = [item for item in scenario.get("states", []) if item.get("state_kind") != "activity_output"]
    for index, state in enumerate(seed_states, 1):
        state_refs[state["id"]] = f"S{index:03d}"
    for activity in scenario.get("activities", []):
        state_refs[activity["output_state_id"]] = f"{activity_refs[activity['id']]}:OUTPUT"

    for item in scenario.get("activity_packages", []):
        workbook["活动包"].append([package_refs[item["id"]], item["name"], package_refs.get(item.get("parent_id"), ""), item.get("sort_order", 0)])
    for item in scenario.get("activities", []):
        workbook["活动"].append([activity_refs[item["id"]], item["name"], item["duration"], item.get("max_instances"), _yn(item["id"] in scenario.get("target_activity_ids", []))])
        for relation in item.get("preconditions", []):
            workbook["活动前置"].append([activity_refs[item["id"]], state_refs.get(relation["state_id"], relation["state_id"]), relation.get("relation_role", "required")])
        for resource_id, quantity in item.get("resource_reqs", {}).items():
            workbook["活动资源"].append([activity_refs[item["id"]], resource_refs.get(resource_id, resource_id), quantity])
        for event_id in item.get("event_reqs", []):
            workbook["活动事件"].append([activity_refs[item["id"]], event_refs.get(event_id, event_id)])
    for item in scenario.get("activity_package_memberships", []):
        workbook["包成员"].append([package_refs[item["package_id"]], activity_refs[item["activity_id"]], item.get("sort_order", 0)])
    for state in seed_states:
        workbook["种子状态"].append([state_refs[state["id"]], state["name"], _yn(state["id"] in scenario.get("initial_state_ids", [])), _yn(state["id"] in scenario.get("goal_state_ids", [])), _yn(state["id"] in scenario.get("forbidden_state_ids", []))])
    for item in scenario.get("resources", []):
        workbook["资源"].append([resource_refs[item["id"]], item.get("name", resource_refs[item["id"]]), item["capacity"]])
    for item in scenario.get("external_events", []):
        workbook["外部事件"].append([event_refs[item["id"]], item.get("name", event_refs[item["id"]]), item["time"]])
        for state_id in item.get("add_state_ids", []): workbook["事件状态"].append([event_refs[item["id"]], "add", state_refs.get(state_id, state_id)])
        for state_id in item.get("remove_state_ids", []): workbook["事件状态"].append([event_refs[item["id"]], "remove", state_refs.get(state_id, state_id)])
    for value in scenario.get("target_activity_package_ids", []): workbook["包目标范围"].append(["target", package_refs[value]])
    for value in scenario.get("activity_package_scope_ids", []): workbook["包目标范围"].append(["scope", package_refs[value]])
    return _bytes(workbook)


def import_workbook(content: bytes) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(content), data_only=True)
    missing = sorted(set(SHEETS) - set(workbook.sheetnames))
    if missing:
        raise PlannerScenarioError("EXCEL_SHEET_MISSING", f"Missing Excel sheets: {', '.join(missing)}")
    rows = {name: _rows(workbook[name]) for name in SHEETS}
    config = rows["场景"][0] if rows["场景"] else {}
    payload: dict[str, Any] = {
        "name": str(config.get("场景名称") or "Excel 导入场景"),
        "execution_mode": str(config.get("执行模式") or "serial"),
        "start_time": int(config.get("起始时间") or 0),
        "max_steps": int(config.get("最大活动实例数") or 20),
        "default_budget": {"time_limit_seconds": float(config.get("时间预算秒") or 5), "transition_limit": int(config.get("转换预算") or 20000), "max_solutions": 20},
        "states": [], "activities": [], "activity_packages": [], "activity_package_memberships": [],
        "resources": [], "external_events": [], "initial_state_ids": [], "goal_state_ids": [],
        "forbidden_state_ids": [], "target_activity_ids": [], "target_activity_package_ids": [],
        "activity_package_scope_ids": [],
    }
    package_refs, activity_refs, state_refs, resource_refs, event_refs = set(), set(), set(), set(), set()
    for row in rows["活动包"]:
        ref = _required(row, "包引用", "活动包")
        if ref in package_refs: raise PlannerScenarioError("EXCEL_REF_DUPLICATE", f"Duplicate package ref: {ref}")
        package_refs.add(ref)
        payload["activity_packages"].append({"id": ref, "display_code": ref, "name": _required(row, "名称", ref), "parent_id": _text(row.get("父包引用")) or None, "level": 2 if _text(row.get("父包引用")) else 1, "sort_order": int(row.get("排序") or 0), "is_active": True})
    for row in rows["种子状态"]:
        ref = _required(row, "状态引用", "种子状态")
        if ref in state_refs: raise PlannerScenarioError("EXCEL_REF_DUPLICATE", f"Duplicate state ref: {ref}")
        state_refs.add(ref)
        payload["states"].append({"id": ref, "name": _required(row, "名称", ref), "state_kind": "seed"})
        if _bool(row.get("初始激活")): payload["initial_state_ids"].append(ref)
        if _bool(row.get("目标必需")): payload["goal_state_ids"].append(ref)
        if _bool(row.get("目标禁止")): payload["forbidden_state_ids"].append(ref)
    activity_by_ref = {}
    for row in rows["活动"]:
        ref = _required(row, "活动引用", "活动")
        if ref in activity_refs: raise PlannerScenarioError("EXCEL_REF_DUPLICATE", f"Duplicate activity ref: {ref}")
        activity_refs.add(ref)
        activity = {"id": ref, "display_code": ref, "name": _required(row, "名称", ref), "duration": int(row.get("工期") or 0), "preconditions": [], "output_state_id": f"{ref}:OUTPUT", "output_state_name": f"{row.get('名称')}完成", "additional_output_state_ids": [], "resource_reqs": {}, "event_reqs": [], "max_instances": int(row["最大实例数"]) if row.get("最大实例数") not in (None, "") else None, "is_active": True}
        payload["activities"].append(activity); activity_by_ref[ref] = activity
        if _bool(row.get("是否目标活动")): payload["target_activity_ids"].append(ref)
    for row in rows["活动前置"]:
        activity_by_ref[_known(row, "活动引用", activity_refs)].setdefault("preconditions", []).append({"state_id": _text(row.get("状态引用")), "relation_role": str(row.get("关系角色") or "required")})
    for row in rows["包成员"]:
        package_ref = _known(row, "包引用", package_refs); activity_ref = _known(row, "活动引用", activity_refs)
        payload["activity_package_memberships"].append({"id": f"MEMBER:{len(payload['activity_package_memberships'])+1}", "package_id": package_ref, "activity_id": activity_ref, "sort_order": int(row.get("排序") or 0), "layout": {}})
    for row in rows["资源"]:
        ref = _required(row, "资源引用", "资源"); resource_refs.add(ref); payload["resources"].append({"id": ref, "name": _required(row, "名称", ref), "capacity": int(row.get("容量") or 0), "is_active": True})
    for row in rows["活动资源"]: activity_by_ref[_known(row, "活动引用", activity_refs)]["resource_reqs"][_known(row, "资源引用", resource_refs)] = int(row.get("需求数量") or 0)
    event_by_ref = {}
    for row in rows["外部事件"]:
        ref = _required(row, "事件引用", "事件"); event_refs.add(ref); event = {"id": ref, "name": _required(row, "名称", ref), "time": int(row.get("发生时间") or 0), "add_state_ids": [], "remove_state_ids": []}; payload["external_events"].append(event); event_by_ref[ref] = event
    for row in rows["事件状态"]:
        event = event_by_ref[_known(row, "事件引用", event_refs)]; operation = str(row.get("操作") or "").lower(); state_ref = _text(row.get("状态引用")); event["add_state_ids" if operation == "add" else "remove_state_ids"].append(state_ref)
    for row in rows["活动事件"]: activity_by_ref[_known(row, "活动引用", activity_refs)]["event_reqs"].append(_known(row, "事件引用", event_refs))
    for row in rows["包目标范围"]:
        package_ref = _known(row, "包引用", package_refs); key = "target_activity_package_ids" if str(row.get("用途") or "target").lower() == "target" else "activity_package_scope_ids"; payload[key].append(package_ref)
    return normalize_import(payload, preserve_ids=False)


def _style(sheet, columns: int) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="0F766E"); cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = f"A1:{chr(64 + columns)}1"
    for column in range(1, columns + 1): sheet.column_dimensions[chr(64 + column)].width = 20


def _bytes(workbook) -> bytes:
    stream = BytesIO(); workbook.save(stream); return stream.getvalue()


def _rows(sheet) -> list[dict[str, Any]]:
    values = list(sheet.iter_rows(values_only=True)); headers = [str(value) if value is not None else "" for value in (values[0] if values else [])]
    return [{headers[index]: value for index, value in enumerate(row) if index < len(headers)} for row in values[1:] if any(value not in (None, "") for value in row)]


def _text(value: Any) -> str: return "" if value is None else str(value).strip()
def _required(row, field, context):
    value = _text(row.get(field))
    if not value: raise PlannerScenarioError("EXCEL_FIELD_REQUIRED", f"{context}: {field} is required")
    return value
def _known(row, field, values):
    value = _required(row, field, field)
    if value not in values: raise PlannerScenarioError("EXCEL_REF_UNKNOWN", f"Unknown {field}: {value}")
    return value
def _bool(value: Any) -> bool: return str(value or "").strip().lower() in {"1", "true", "yes", "y", "是"}
def _yn(value: Any) -> str: return "是" if value else "否"
