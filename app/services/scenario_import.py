"""Scenario Excel import parsing, validation, and persistence.

This module keeps bulk import concerns outside the CRUD routes and outside the
planner/scheduler pipeline. It parses a workbook into row dictionaries, validates
cross-sheet references, estimates create/update counts, and applies strict
upsert writes when requested.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.db.models import (
    ActivityPackageAtomicRef,
    ActivityNode,
    AtomicActivity,
    FeatureDefinition,
    Machine,
    MachineState,
    MachineStateFeature,
    MachineType,
    MaintenanceIntentTemplate,
    OpRule,
    OpRuleEffect,
    OpRulePrecond,
    OpRuleResourceReq,
    Resource,
    ScopeGuard,
    ScopeGuardPrecond,
    StateFeatureDef,
    StateNode,
)
from app.db.schemas import LayeredExpansionRequest
from app.services.layered_health import check_layered_health


REQUIRED_SHEETS = [
    "meta",
    "feature_catalog",
    "machine_type",
    "machines",
    "state_feature_defs",
    "resources",
    "rules",
    "states",
    "solve_cases",
    "instructions",
]

OPTIONAL_SHEETS = [
    "activity_nodes",
    "atomic_activities",
    "activity_package_atomic_refs",
    "state_nodes",
    "scope_guards",
    "maintenance_intents",
    "layered_health_checks",
    "rule_groups",
    "notes",
]

SHEET_FIELDS = {
    "meta": ["scenario_code", "scenario_name", "version", "mode"],
    "feature_catalog": ["feature_key", "value_type", "allowed_values", "unit", "description"],
    "machine_type": ["code", "name", "description"],
    "machines": ["code", "machine_type_code", "name", "location"],
    "state_feature_defs": ["machine_type_code", "feature_key", "feature_name", "value_type", "allowed_values"],
    "resources": ["machine_code", "code", "name", "resource_type", "capacity", "is_available", "meta_json"],
    "rules": [
        "code",
        "machine_type_code",
        "name",
        "duration_min",
        "description",
        "is_active",
        "is_repair",
        "preconditions",
        "effects",
        "resource_reqs",
    ],
    "states": ["machine_code", "state_code", "state_type", "label", "features"],
    "solve_cases": [
        "case_code",
        "machine_code",
        "current_state_code",
        "target_state_code",
        "objective",
        "objectives_json",
        "constraints_json",
        "expected_min_steps",
        "expected_max_makespan_min",
    ],
    "activity_nodes": [
        "machine_type_code",
        "code",
        "parent_code",
        "level",
        "name",
        "activity_category",
        "sort_order",
        "is_active",
        "metadata_json",
    ],
    "atomic_activities": [
        "machine_type_code",
        "code",
        "name",
        "activity_category",
        "sort_order",
        "is_active",
        "metadata_json",
    ],
    "activity_package_atomic_refs": [
        "machine_type_code",
        "package_code",
        "atomic_activity_code",
        "sort_order",
        "is_active",
        "metadata_json",
    ],
    "state_nodes": [
        "machine_type_code",
        "code",
        "parent_code",
        "level",
        "name",
        "feature_key",
        "operator",
        "target_value",
        "state_kind",
        "sort_order",
        "is_active",
        "metadata_json",
    ],
    "scope_guards": [
        "machine_type_code",
        "activity_node_code",
        "name",
        "description",
        "is_active",
        "preconditions",
        "metadata_json",
    ],
    "maintenance_intents": [
        "machine_type_code",
        "issue_type",
        "name",
        "scope_activity_node_code",
        "description",
        "target_state_node_codes",
        "candidate_activity_scope_codes",
        "observed_fact_templates",
        "desired_fact_templates",
        "is_active",
        "metadata_json",
    ],
    "layered_health_checks": [
        "machine_type_code",
        "check_code",
        "name",
        "target_state_node_codes",
        "activity_scope_node_codes",
        "include_inactive",
        "description",
    ],
}

VALID_VALUE_TYPES = {"string", "number", "boolean", "enum"}
VALID_OPERATORS = {"eq", "neq", "gt", "gte", "lt", "lte", "in"}
VALID_STATE_NODE_OPERATORS = VALID_OPERATORS | {"completed"}
VALID_EFFECT_TYPES = {"set", "increment", "decrement", "sub", "reset"}
VALID_STATE_TYPES = {"current", "target", "snapshot"}
VALID_STATE_KINDS = {"aggregate", "atomic", "external", "manual"}


@dataclass
class ImportErrorItem:
    sheet: str
    row: int | None
    field: str | None
    message: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "row": self.row,
            "field": self.field,
            "message": self.message,
        }


@dataclass
class ParsedRow:
    sheet: str
    row_number: int
    data: dict[str, Any]


@dataclass
class ScenarioWorkbook:
    rows: dict[str, list[ParsedRow]] = field(default_factory=dict)
    errors: list[ImportErrorItem] = field(default_factory=list)

    def sheet(self, name: str) -> list[ParsedRow]:
        return self.rows.get(name, [])


def _blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _normalize_header(value: Any) -> str:
    return _text(value).lower()


def _parse_allowed_values(value: Any) -> list[str] | None:
    if _blank(value):
        return None
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in str(value).split(",") if item.strip()]


def _parse_bool(value: Any, *, default: bool | None = None) -> bool | None:
    if _blank(value):
        return default
    lowered = str(value).strip().lower()
    if lowered in {"true", "1", "yes", "y", "是"}:
        return True
    if lowered in {"false", "0", "no", "n", "否"}:
        return False
    return None


def _parse_int(value: Any) -> int | None:
    if _blank(value):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _parse_json(value: Any) -> Any:
    if _blank(value):
        return None
    if isinstance(value, (dict, list)):
        return value
    return json.loads(str(value))


def _parse_preconditions(raw: Any, row: ParsedRow, errors: list[ImportErrorItem]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    if _blank(raw):
        return items
    for token in str(raw).split(";"):
        token = token.strip()
        if not token:
            continue
        parts = [part.strip() for part in token.split(":", 2)]
        if len(parts) != 3:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "preconditions", f"Invalid item: {token}"))
            continue
        feature_key, operator, feature_value = parts
        if operator not in VALID_OPERATORS:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "preconditions", f"Invalid operator: {operator}"))
            continue
        item: dict[str, Any] = {
            "feature_key": feature_key,
            "operator": operator,
            "feature_value": feature_value,
            "value_list": None,
        }
        if operator == "in":
            item["value_list"] = [value.strip() for value in feature_value.split(",") if value.strip()]
        items.append(item)
    return items


def _parse_effects(raw: Any, row: ParsedRow, errors: list[ImportErrorItem]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    if _blank(raw):
        return items
    for token in str(raw).split(";"):
        token = token.strip()
        if not token:
            continue
        parts = [part.strip() for part in token.split(":", 2)]
        if len(parts) != 3:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "effects", f"Invalid item: {token}"))
            continue
        feature_key, effect_type, value = parts
        if effect_type not in VALID_EFFECT_TYPES:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "effects", f"Invalid effect_type: {effect_type}"))
            continue
        item: dict[str, Any] = {
            "feature_key": feature_key,
            "effect_type": effect_type,
            "new_value": value,
            "delta_value": None,
        }
        if effect_type in {"increment", "decrement", "sub"}:
            try:
                item["delta_value"] = float(value)
            except ValueError:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "effects", f"Invalid numeric delta: {value}"))
                continue
        items.append(item)
    return items


def _parse_resource_reqs(raw: Any, row: ParsedRow, errors: list[ImportErrorItem]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    if _blank(raw):
        return items
    for token in str(raw).split(";"):
        token = token.strip()
        if not token:
            continue
        parts = [part.strip() for part in token.split(":", 2)]
        if len(parts) != 3:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "resource_reqs", f"Invalid item: {token}"))
            continue
        resource_type, quantity_raw, required_raw = parts
        quantity = _parse_int(quantity_raw)
        is_required = _parse_bool(required_raw)
        if quantity is None or quantity < 1:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "resource_reqs", f"Invalid quantity: {quantity_raw}"))
            continue
        if is_required is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "resource_reqs", f"Invalid is_required: {required_raw}"))
            continue
        items.append({"resource_type": resource_type, "quantity": quantity, "is_required": is_required})
    return items


def _parse_state_features(raw: Any, row: ParsedRow, errors: list[ImportErrorItem]) -> dict[str, str]:
    items: dict[str, str] = {}
    if _blank(raw):
        return items
    for token in str(raw).split(";"):
        token = token.strip()
        if not token:
            continue
        parts = [part.strip() for part in token.split(":", 1)]
        if len(parts) != 2:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "features", f"Invalid item: {token}"))
            continue
        items[parts[0]] = parts[1]
    return items


def _parse_scope_guard_preconditions(
    raw: Any,
    row: ParsedRow,
    errors: list[ImportErrorItem],
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    if _blank(raw):
        return items
    for token in str(raw).split(";"):
        token = token.strip()
        if not token:
            continue
        parts = [part.strip() for part in token.split(":", 2)]
        if len(parts) not in (2, 3):
            errors.append(ImportErrorItem(row.sheet, row.row_number, "preconditions", f"Invalid item: {token}"))
            continue
        state_node_code = parts[0]
        operator = parts[1] or "completed"
        expected_value = parts[2] if len(parts) == 3 else None
        if operator not in VALID_STATE_NODE_OPERATORS:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "preconditions", f"Invalid operator: {operator}"))
            continue
        if operator != "completed" and _blank(expected_value):
            errors.append(ImportErrorItem(row.sheet, row.row_number, "preconditions", f"Missing expected value: {token}"))
            continue
        value_list = None
        if operator == "in" and expected_value is not None:
            value_list = [value.strip() for value in expected_value.split(",") if value.strip()]
        items.append(
            {
                "state_node_code": state_node_code,
                "operator": operator,
                "expected_value": expected_value,
                "value_list": value_list,
            }
        )
    return items


def _parse_code_list(raw: Any) -> list[str]:
    if _blank(raw):
        return []
    if isinstance(raw, list):
        return [str(item).strip() for item in raw if str(item).strip()]
    normalized = str(raw).replace("\n", ";").replace(",", ";")
    return [token.strip() for token in normalized.split(";") if token.strip()]


def _parse_maintenance_fact_templates(
    raw: Any,
    row: ParsedRow,
    field_name: str,
    errors: list[ImportErrorItem],
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    if _blank(raw):
        return items

    parsed_json: Any | None = None
    if isinstance(raw, (dict, list)):
        parsed_json = raw
    else:
        raw_text = str(raw).strip()
        if raw_text.startswith(("[", "{")):
            try:
                parsed_json = _parse_json(raw_text)
            except json.JSONDecodeError:
                errors.append(ImportErrorItem(row.sheet, row.row_number, field_name, "Invalid JSON"))
                return items

    if parsed_json is not None:
        parsed_items = parsed_json if isinstance(parsed_json, list) else [parsed_json]
        for item in parsed_items:
            if not isinstance(item, dict):
                errors.append(ImportErrorItem(row.sheet, row.row_number, field_name, "Fact template items must be objects"))
                continue
            feature_key = _text(item.get("feature_key"))
            operator = _text(item.get("operator")) or "eq"
            value = item.get("value")
            value_list = item.get("value_list")
            if not feature_key:
                errors.append(ImportErrorItem(row.sheet, row.row_number, field_name, "Fact template feature_key is required"))
                continue
            if value is None or _blank(value):
                errors.append(ImportErrorItem(row.sheet, row.row_number, field_name, f"Fact '{feature_key}' requires a value"))
                continue
            items.append(
                {
                    "feature_key": feature_key,
                    "operator": operator,
                    "value": str(value),
                    "value_list": value_list,
                }
            )
        return items

    for token in str(raw).split(";"):
        token = token.strip()
        if not token:
            continue
        parts = [part.strip() for part in token.split(":", 2)]
        if len(parts) == 2:
            feature_key, value = parts
            operator = "eq"
        elif len(parts) == 3:
            feature_key, operator, value = parts
        else:
            errors.append(ImportErrorItem(row.sheet, row.row_number, field_name, f"Invalid item: {token}"))
            continue
        if not feature_key:
            errors.append(ImportErrorItem(row.sheet, row.row_number, field_name, f"Missing feature_key: {token}"))
            continue
        if operator != "eq":
            errors.append(
                ImportErrorItem(row.sheet, row.row_number, field_name, "Maintenance fact templates support operator 'eq' only")
            )
            continue
        if _blank(value):
            errors.append(ImportErrorItem(row.sheet, row.row_number, field_name, f"Missing value: {token}"))
            continue
        items.append(
            {
                "feature_key": feature_key,
                "operator": operator,
                "value": value,
                "value_list": None,
            }
        )
    return items


def parse_scenario_workbook(content: bytes) -> ScenarioWorkbook:
    parsed = ScenarioWorkbook()
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        parsed.errors.append(ImportErrorItem("workbook", None, None, f"Unable to read .xlsx file: {exc}"))
        return parsed

    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in workbook.sheetnames:
            parsed.errors.append(ImportErrorItem(sheet_name, None, None, "Missing required sheet"))

    for sheet_name in [*REQUIRED_SHEETS, *OPTIONAL_SHEETS]:
        if sheet_name == "instructions" or sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [_normalize_header(value) for value in header_row or []]
        expected = SHEET_FIELDS.get(sheet_name, [])
        for field_name in expected:
            if field_name not in headers:
                parsed.errors.append(ImportErrorItem(sheet_name, 1, field_name, "Missing required column"))
        rows: list[ParsedRow] = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header}
            if not row_data or all(_blank(value) for value in row_data.values()):
                continue
            rows.append(ParsedRow(sheet_name, row_number, row_data))
        parsed.rows[sheet_name] = rows

    return parsed


def _required(row: ParsedRow, field_name: str, errors: list[ImportErrorItem]) -> str:
    value = _text(row.data.get(field_name))
    if not value:
        errors.append(ImportErrorItem(row.sheet, row.row_number, field_name, "Required field is empty"))
    return value


def _check_unique(rows: list[ParsedRow], fields: list[str], errors: list[ImportErrorItem]) -> None:
    seen: dict[tuple[str, ...], ParsedRow] = {}
    for row in rows:
        key = tuple(_text(row.data.get(field_name)) for field_name in fields)
        if any(not item for item in key):
            continue
        if key in seen:
            errors.append(
                ImportErrorItem(row.sheet, row.row_number, ",".join(fields), f"Duplicate business key: {' / '.join(key)}")
            )
        else:
            seen[key] = row


async def _load_existing(session: AsyncSession) -> dict[str, Any]:
    machine_types = (await session.execute(select(MachineType))).scalars().all()
    machines = (await session.execute(select(Machine).options(selectinload(Machine.machine_type)))).scalars().all()
    features = (await session.execute(select(FeatureDefinition))).scalars().all()
    state_defs = (await session.execute(select(StateFeatureDef).join(MachineType))).scalars().all()
    resources = (await session.execute(select(Resource).options(selectinload(Resource.machine)))).scalars().all()
    rules = (await session.execute(select(OpRule))).scalars().all()
    activity_nodes = (
        await session.execute(
            select(ActivityNode).options(
                selectinload(ActivityNode.machine_type),
                selectinload(ActivityNode.parent),
            )
        )
    ).scalars().all()
    atomic_activities = (
        await session.execute(
            select(AtomicActivity).options(selectinload(AtomicActivity.machine_type))
        )
    ).scalars().all()
    atomic_refs = (
        await session.execute(
            select(ActivityPackageAtomicRef).options(
                selectinload(ActivityPackageAtomicRef.activity_node).selectinload(ActivityNode.machine_type),
                selectinload(ActivityPackageAtomicRef.atomic_activity),
            )
        )
    ).scalars().all()
    state_nodes = (
        await session.execute(
            select(StateNode).options(
                selectinload(StateNode.machine_type),
                selectinload(StateNode.parent),
            )
        )
    ).scalars().all()
    scope_guards = (
        await session.execute(
            select(ScopeGuard).options(
                selectinload(ScopeGuard.activity_node).selectinload(ActivityNode.machine_type)
            )
        )
    ).scalars().all()
    maintenance_intents = (
        await session.execute(
            select(MaintenanceIntentTemplate).options(selectinload(MaintenanceIntentTemplate.machine_type))
        )
    ).scalars().all()

    machine_type_by_id = {item.id: item for item in machine_types}
    state_defs_by_type: dict[str, set[str]] = {}
    for item in state_defs:
        machine_type = machine_type_by_id.get(item.machine_type_id)
        if machine_type:
            state_defs_by_type.setdefault(machine_type.code, set()).add(item.feature_key)

    return {
        "machine_types": {item.code: item for item in machine_types},
        "machines": {item.code: item for item in machines},
        "features": {item.feature_key: item for item in features},
        "resources": {
            (item.machine.code, item.code): item
            for item in resources
            if item.machine is not None
        },
        "resource_types": {item.resource_type for item in resources},
        "rules": {item.code: item for item in rules},
        "state_defs_by_type": state_defs_by_type,
        "activity_nodes": {
            (item.machine_type.code, item.code): item
            for item in activity_nodes
            if item.machine_type is not None
        },
        "atomic_activities": {
            (item.machine_type.code, item.code): item
            for item in atomic_activities
            if item.machine_type is not None
        },
        "activity_package_atomic_refs": {
            (item.activity_node.machine_type.code, item.activity_node.code, item.atomic_activity.code): item
            for item in atomic_refs
            if item.activity_node is not None
            and item.activity_node.machine_type is not None
            and item.atomic_activity is not None
        },
        "state_nodes": {
            (item.machine_type.code, item.code): item
            for item in state_nodes
            if item.machine_type is not None
        },
        "scope_guards": {
            (item.activity_node.machine_type.code, item.activity_node.code, item.name): item
            for item in scope_guards
            if item.activity_node is not None and item.activity_node.machine_type is not None
        },
        "maintenance_intents": {
            (item.machine_type.code, item.issue_type): item
            for item in maintenance_intents
            if item.machine_type is not None
        },
    }


async def validate_scenario_workbook(parsed: ScenarioWorkbook, session: AsyncSession) -> dict[str, Any]:
    errors = parsed.errors
    existing = await _load_existing(session)

    meta_rows = parsed.sheet("meta")
    if len(meta_rows) != 1:
        errors.append(ImportErrorItem("meta", None, None, "meta sheet must contain exactly one data row"))
    scenario_code = _text(meta_rows[0].data.get("scenario_code")) if meta_rows else ""
    scenario_name = _text(meta_rows[0].data.get("scenario_name")) if meta_rows else ""
    mode = _text(meta_rows[0].data.get("mode")) if meta_rows else ""
    if meta_rows:
        _required(meta_rows[0], "scenario_code", errors)
        _required(meta_rows[0], "scenario_name", errors)
        if mode != "scenario_upsert":
            errors.append(ImportErrorItem("meta", meta_rows[0].row_number, "mode", "mode must be scenario_upsert"))

    for sheet_name, fields in {
        "feature_catalog": ["feature_key"],
        "machine_type": ["code"],
        "machines": ["code"],
        "state_feature_defs": ["machine_type_code", "feature_key"],
        "resources": ["machine_code", "code"],
        "activity_nodes": ["machine_type_code", "code"],
        "atomic_activities": ["machine_type_code", "code"],
        "activity_package_atomic_refs": ["machine_type_code", "package_code", "atomic_activity_code"],
        "state_nodes": ["machine_type_code", "code"],
        "scope_guards": ["machine_type_code", "activity_node_code", "name"],
        "maintenance_intents": ["machine_type_code", "issue_type"],
        "layered_health_checks": ["machine_type_code", "check_code"],
        "rules": ["code"],
        "states": ["machine_code", "state_code"],
        "solve_cases": ["case_code"],
    }.items():
        _check_unique(parsed.sheet(sheet_name), fields, errors)

    feature_rows = parsed.sheet("feature_catalog")
    machine_type_rows = parsed.sheet("machine_type")
    machine_rows = parsed.sheet("machines")
    state_def_rows = parsed.sheet("state_feature_defs")
    resource_rows = parsed.sheet("resources")
    activity_node_rows = parsed.sheet("activity_nodes")
    atomic_activity_rows = parsed.sheet("atomic_activities")
    atomic_ref_rows = parsed.sheet("activity_package_atomic_refs")
    state_node_rows = parsed.sheet("state_nodes")
    scope_guard_rows = parsed.sheet("scope_guards")
    maintenance_intent_rows = parsed.sheet("maintenance_intents")
    layered_health_check_rows = parsed.sheet("layered_health_checks")
    rule_rows = parsed.sheet("rules")
    state_rows = parsed.sheet("states")
    solve_case_rows = parsed.sheet("solve_cases")

    file_features: set[str] = set()
    for row in feature_rows:
        feature_key = _required(row, "feature_key", errors)
        value_type = _required(row, "value_type", errors)
        if value_type and value_type not in VALID_VALUE_TYPES:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "value_type", f"Invalid value_type: {value_type}"))
        file_features.add(feature_key)

    machine_type_codes = set(existing["machine_types"].keys())
    for row in machine_type_rows:
        code = _required(row, "code", errors)
        _required(row, "name", errors)
        machine_type_codes.add(code)

    machine_to_type: dict[str, str] = {
        code: machine.machine_type.code for code, machine in existing["machines"].items() if machine.machine_type
    }
    for row in machine_rows:
        code = _required(row, "code", errors)
        machine_type_code = _required(row, "machine_type_code", errors)
        _required(row, "name", errors)
        if machine_type_code and machine_type_code not in machine_type_codes:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "machine_type_code", f"Unknown machine_type_code: {machine_type_code}"))
        machine_to_type[code] = machine_type_code

    feature_catalog = set(existing["features"].keys()) | file_features
    state_defs_by_type: dict[str, set[str]] = {
        code: set(values) for code, values in existing["state_defs_by_type"].items()
    }
    for row in state_def_rows:
        machine_type_code = _required(row, "machine_type_code", errors)
        feature_key = _required(row, "feature_key", errors)
        value_type = _required(row, "value_type", errors)
        if machine_type_code and machine_type_code not in machine_type_codes:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "machine_type_code", f"Unknown machine_type_code: {machine_type_code}"))
        if feature_key and feature_key not in feature_catalog:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "feature_key", f"Unknown feature_key: {feature_key}"))
        if value_type and value_type not in VALID_VALUE_TYPES:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "value_type", f"Invalid value_type: {value_type}"))
        state_defs_by_type.setdefault(machine_type_code, set()).add(feature_key)

    activity_node_info: dict[tuple[str, str], dict[str, Any]] = {
        key: {
            "level": node.level,
            "machine_type_code": key[0],
            "parent_code": node.parent.code if node.parent else "",
        }
        for key, node in existing["activity_nodes"].items()
    }
    for row in activity_node_rows:
        machine_type_code = _required(row, "machine_type_code", errors)
        code = _required(row, "code", errors)
        _required(row, "name", errors)
        level = _parse_int(row.data.get("level"))
        if machine_type_code and machine_type_code not in machine_type_codes:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "machine_type_code", f"Unknown machine_type_code: {machine_type_code}"))
        if level not in (1, 2, 3):
            errors.append(ImportErrorItem(row.sheet, row.row_number, "level", "level must be 1, 2, or 3"))
            level = 0
        active = _parse_bool(row.data.get("is_active"), default=True)
        if active is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "is_active", "Invalid boolean value"))
        if _parse_int(row.data.get("sort_order")) is None and not _blank(row.data.get("sort_order")):
            errors.append(ImportErrorItem(row.sheet, row.row_number, "sort_order", "sort_order must be an integer"))
        if not _blank(row.data.get("metadata_json")):
            try:
                _parse_json(row.data.get("metadata_json"))
            except json.JSONDecodeError:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "metadata_json", "Invalid JSON object"))
        activity_node_info[(machine_type_code, code)] = {
            "level": level,
            "machine_type_code": machine_type_code,
            "parent_code": _text(row.data.get("parent_code")),
        }

    for row in activity_node_rows:
        machine_type_code = _text(row.data.get("machine_type_code"))
        code = _text(row.data.get("code"))
        info = activity_node_info.get((machine_type_code, code), {})
        level = int(info.get("level") or 0)
        parent_code = _text(row.data.get("parent_code"))
        if level == 1 and parent_code:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "parent_code", "Level-1 activity nodes cannot have a parent"))
        if level in (2, 3) and not parent_code:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "parent_code", "Level-2/3 activity nodes require a parent"))
        if parent_code:
            parent_info = activity_node_info.get((machine_type_code, parent_code))
            if parent_info is None:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "parent_code", f"Unknown parent activity node: {parent_code}"))
            elif parent_info.get("level") != level - 1:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "parent_code", "Activity parent level must be exactly one level above child"))

    atomic_activity_info: dict[tuple[str, str], dict[str, Any]] = {
        key: {"machine_type_code": key[0]}
        for key in existing["atomic_activities"]
    }
    for (machine_type_code, code), info in activity_node_info.items():
        if info.get("level") == 3:
            atomic_activity_info.setdefault((machine_type_code, code), {"machine_type_code": machine_type_code})

    for row in atomic_activity_rows:
        machine_type_code = _required(row, "machine_type_code", errors)
        code = _required(row, "code", errors)
        _required(row, "name", errors)
        if machine_type_code and machine_type_code not in machine_type_codes:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "machine_type_code", f"Unknown machine_type_code: {machine_type_code}"))
        active = _parse_bool(row.data.get("is_active"), default=True)
        if active is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "is_active", "Invalid boolean value"))
        if _parse_int(row.data.get("sort_order")) is None and not _blank(row.data.get("sort_order")):
            errors.append(ImportErrorItem(row.sheet, row.row_number, "sort_order", "sort_order must be an integer"))
        if not _blank(row.data.get("metadata_json")):
            try:
                _parse_json(row.data.get("metadata_json"))
            except json.JSONDecodeError:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "metadata_json", "Invalid JSON object"))
        atomic_activity_info[(machine_type_code, code)] = {"machine_type_code": machine_type_code}

    for row in atomic_ref_rows:
        machine_type_code = _required(row, "machine_type_code", errors)
        package_code = _required(row, "package_code", errors)
        atomic_activity_code = _required(row, "atomic_activity_code", errors)
        package_info = activity_node_info.get((machine_type_code, package_code))
        if package_info is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "package_code", f"Unknown activity package: {package_code}"))
        elif package_info.get("level") != 2:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "package_code", "Atomic activity refs can only attach to level-2 activity packages"))
        if (machine_type_code, atomic_activity_code) not in atomic_activity_info:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "atomic_activity_code", f"Unknown atomic activity: {atomic_activity_code}"))
        active = _parse_bool(row.data.get("is_active"), default=True)
        if active is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "is_active", "Invalid boolean value"))
        if _parse_int(row.data.get("sort_order")) is None and not _blank(row.data.get("sort_order")):
            errors.append(ImportErrorItem(row.sheet, row.row_number, "sort_order", "sort_order must be an integer"))
        if not _blank(row.data.get("metadata_json")):
            try:
                _parse_json(row.data.get("metadata_json"))
            except json.JSONDecodeError:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "metadata_json", "Invalid JSON object"))

    state_node_info: dict[tuple[str, str], dict[str, Any]] = {
        key: {
            "level": node.level,
            "machine_type_code": key[0],
            "parent_code": node.parent.code if node.parent else "",
        }
        for key, node in existing["state_nodes"].items()
    }
    for row in state_node_rows:
        machine_type_code = _required(row, "machine_type_code", errors)
        code = _required(row, "code", errors)
        _required(row, "name", errors)
        level = _parse_int(row.data.get("level"))
        if machine_type_code and machine_type_code not in machine_type_codes:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "machine_type_code", f"Unknown machine_type_code: {machine_type_code}"))
        if level is None or level < 1:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "level", "level must be a positive integer"))
            level = 0
        operator = _text(row.data.get("operator")) or "eq"
        if operator not in VALID_STATE_NODE_OPERATORS:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "operator", f"Invalid operator: {operator}"))
        feature_key = _text(row.data.get("feature_key"))
        target_value = _text(row.data.get("target_value"))
        state_kind = _text(row.data.get("state_kind")) or ("aggregate" if not feature_key and not target_value else "atomic")
        if state_kind not in VALID_STATE_KINDS:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "state_kind", f"Invalid state_kind: {state_kind}"))
        if state_kind == "aggregate":
            if feature_key or target_value:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "feature_key", "Aggregate state nodes cannot bind feature values"))
        else:
            if not feature_key:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "feature_key", "Atomic state nodes require a feature_key"))
            else:
                feature_catalog.add(feature_key)
                state_defs_by_type.setdefault(machine_type_code, set()).add(feature_key)
            if operator != "eq":
                errors.append(ImportErrorItem(row.sheet, row.row_number, "operator", "Atomic state nodes currently support operator 'eq' only"))
            if not target_value:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "target_value", "Atomic state nodes require a target_value"))
        active = _parse_bool(row.data.get("is_active"), default=True)
        if active is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "is_active", "Invalid boolean value"))
        if _parse_int(row.data.get("sort_order")) is None and not _blank(row.data.get("sort_order")):
            errors.append(ImportErrorItem(row.sheet, row.row_number, "sort_order", "sort_order must be an integer"))
        if not _blank(row.data.get("metadata_json")):
            try:
                _parse_json(row.data.get("metadata_json"))
            except json.JSONDecodeError:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "metadata_json", "Invalid JSON object"))
        state_node_info[(machine_type_code, code)] = {
            "level": level,
            "machine_type_code": machine_type_code,
            "parent_code": _text(row.data.get("parent_code")),
        }

    for row in state_node_rows:
        machine_type_code = _text(row.data.get("machine_type_code"))
        code = _text(row.data.get("code"))
        info = state_node_info.get((machine_type_code, code), {})
        level = int(info.get("level") or 0)
        parent_code = _text(row.data.get("parent_code"))
        if level == 1 and parent_code:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "parent_code", "Level-1 state nodes cannot have a parent"))
        if level > 1 and not parent_code:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "parent_code", "Non-root state nodes require a parent"))
        if parent_code:
            parent_info = state_node_info.get((machine_type_code, parent_code))
            if parent_info is None:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "parent_code", f"Unknown parent state node: {parent_code}"))
            elif parent_info.get("level") != level - 1:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "parent_code", "State parent level must be exactly one level above child"))

    for row in scope_guard_rows:
        machine_type_code = _required(row, "machine_type_code", errors)
        activity_node_code = _required(row, "activity_node_code", errors)
        _required(row, "name", errors)
        if machine_type_code and machine_type_code not in machine_type_codes:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "machine_type_code", f"Unknown machine_type_code: {machine_type_code}"))
        activity_info = activity_node_info.get((machine_type_code, activity_node_code))
        if activity_info is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "activity_node_code", f"Unknown activity_node_code: {activity_node_code}"))
            activity_level = None
        else:
            activity_level = activity_info.get("level")
            if activity_level not in (1, 2):
                errors.append(ImportErrorItem(row.sheet, row.row_number, "activity_node_code", "Scope Guards can only attach to level-1 or level-2 activity nodes"))
        active = _parse_bool(row.data.get("is_active"), default=True)
        if active is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "is_active", "Invalid boolean value"))
        if not _blank(row.data.get("metadata_json")):
            try:
                _parse_json(row.data.get("metadata_json"))
            except json.JSONDecodeError:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "metadata_json", "Invalid JSON object"))
        preconditions = _parse_scope_guard_preconditions(row.data.get("preconditions"), row, errors)
        if not preconditions:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "preconditions", "Scope Guard must contain at least one precondition"))
        for item in preconditions:
            state_code = item["state_node_code"]
            state_info = state_node_info.get((machine_type_code, state_code))
            if state_info is None:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "preconditions", f"Unknown state_node_code: {state_code}"))
            elif activity_level == 1 and state_info.get("level") != 1:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "preconditions", "Level-1 activity Scope Guards can only reference level-1 state nodes"))

    for row in maintenance_intent_rows:
        machine_type_code = _required(row, "machine_type_code", errors)
        issue_type = _required(row, "issue_type", errors)
        _required(row, "name", errors)
        scope_activity_node_code = _required(row, "scope_activity_node_code", errors)
        if machine_type_code and machine_type_code not in machine_type_codes:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "machine_type_code", f"Unknown machine_type_code: {machine_type_code}"))
        scope_info = activity_node_info.get((machine_type_code, scope_activity_node_code))
        if scope_info is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "scope_activity_node_code", f"Unknown activity node: {scope_activity_node_code}"))
        elif scope_info.get("level") != 2:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "scope_activity_node_code", "Maintenance intent scope must be a level-2 activity node"))
        target_state_node_codes = _parse_code_list(row.data.get("target_state_node_codes"))
        for state_node_code in target_state_node_codes:
            if (machine_type_code, state_node_code) not in state_node_info:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "target_state_node_codes", f"Unknown state node: {state_node_code}"))
        candidate_activity_scope_codes = _parse_code_list(row.data.get("candidate_activity_scope_codes"))
        for activity_node_code in candidate_activity_scope_codes:
            if (machine_type_code, activity_node_code) not in activity_node_info:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "candidate_activity_scope_codes", f"Unknown activity node: {activity_node_code}"))
        observed_facts = _parse_maintenance_fact_templates(
            row.data.get("observed_fact_templates"),
            row,
            "observed_fact_templates",
            errors,
        )
        desired_facts = _parse_maintenance_fact_templates(
            row.data.get("desired_fact_templates"),
            row,
            "desired_fact_templates",
            errors,
        )
        allowed_features = state_defs_by_type.get(machine_type_code, set())
        for field_name, facts in (
            ("observed_fact_templates", observed_facts),
            ("desired_fact_templates", desired_facts),
        ):
            for item in facts:
                if item["feature_key"] not in allowed_features:
                    errors.append(ImportErrorItem(row.sheet, row.row_number, field_name, f"Unknown feature_key for machine type: {item['feature_key']}"))
                if item["operator"] != "eq":
                    errors.append(ImportErrorItem(row.sheet, row.row_number, field_name, "Maintenance fact templates support operator 'eq' only"))
                if _blank(item["value"]):
                    errors.append(ImportErrorItem(row.sheet, row.row_number, field_name, f"Missing value for fact: {item['feature_key']}"))
        if not target_state_node_codes and not desired_facts:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "target_state_node_codes", "Maintenance intent requires at least one target state node or desired fact"))
        active = _parse_bool(row.data.get("is_active"), default=True)
        if active is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "is_active", "Invalid boolean value"))
        if not _blank(row.data.get("metadata_json")):
            try:
                _parse_json(row.data.get("metadata_json"))
            except json.JSONDecodeError:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "metadata_json", "Invalid JSON object"))
        if issue_type and len(issue_type) > 64:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "issue_type", "issue_type must be 64 characters or fewer"))

    for row in layered_health_check_rows:
        machine_type_code = _required(row, "machine_type_code", errors)
        check_code = _required(row, "check_code", errors)
        _required(row, "name", errors)
        if machine_type_code and machine_type_code not in machine_type_codes:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "machine_type_code", f"Unknown machine_type_code: {machine_type_code}"))
        target_state_node_codes = _parse_code_list(row.data.get("target_state_node_codes"))
        if not target_state_node_codes:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "target_state_node_codes", "At least one target state node code is required"))
        for state_node_code in target_state_node_codes:
            if (machine_type_code, state_node_code) not in state_node_info:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "target_state_node_codes", f"Unknown state node: {state_node_code}"))
        activity_scope_node_codes = _parse_code_list(row.data.get("activity_scope_node_codes"))
        if not activity_scope_node_codes:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "activity_scope_node_codes", "At least one activity scope node code is required"))
        for activity_node_code in activity_scope_node_codes:
            if (machine_type_code, activity_node_code) not in activity_node_info:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "activity_scope_node_codes", f"Unknown activity node: {activity_node_code}"))
        include_inactive = _parse_bool(row.data.get("include_inactive"), default=False)
        if include_inactive is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "include_inactive", "Invalid boolean value"))
        if check_code and len(check_code) > 64:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "check_code", "check_code must be 64 characters or fewer"))

    resource_types_by_machine_type: dict[str, set[str]] = {}
    resource_capacity_by_machine_type: dict[tuple[str, str], int] = {}
    resource_capacity_by_machine_and_type: dict[tuple[str, str], int] = {}
    for (machine_code, _code), resource in existing["resources"].items():
        machine_type_code = machine_to_type.get(machine_code)
        if not machine_type_code:
            continue
        resource_types_by_machine_type.setdefault(machine_type_code, set()).add(resource.resource_type)
        if resource.is_available:
            machine_key = (machine_code, resource.resource_type)
            resource_capacity_by_machine_and_type[machine_key] = (
                resource_capacity_by_machine_and_type.get(machine_key, 0) + resource.capacity
            )
    for (machine_code, resource_type), capacity in resource_capacity_by_machine_and_type.items():
        machine_type_code = machine_to_type.get(machine_code)
        if machine_type_code:
            key = (machine_type_code, resource_type)
            resource_capacity_by_machine_type[key] = max(resource_capacity_by_machine_type.get(key, 0), capacity)

    for row in resource_rows:
        machine_code = _required(row, "machine_code", errors)
        code = _required(row, "code", errors)
        resource_type = _required(row, "resource_type", errors)
        _required(row, "name", errors)
        if machine_code and machine_code not in machine_to_type:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "machine_code", f"Unknown machine_code: {machine_code}"))
        capacity = _parse_int(row.data.get("capacity"))
        if capacity is None or capacity < 1:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "capacity", "capacity must be a positive integer"))
            capacity = 0
        available = _parse_bool(row.data.get("is_available"), default=True)
        if available is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "is_available", "Invalid boolean value"))
            available = True
        if not _blank(row.data.get("meta_json")):
            try:
                _parse_json(row.data.get("meta_json"))
            except json.JSONDecodeError:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "meta_json", "Invalid JSON object"))
        machine_type_code = machine_to_type.get(machine_code)
        if machine_type_code:
            resource_types_by_machine_type.setdefault(machine_type_code, set()).add(resource_type)
            if available:
                machine_key = (machine_code, resource_type)
                resource_capacity_by_machine_and_type[machine_key] = (
                    resource_capacity_by_machine_and_type.get(machine_key, 0) + capacity
                )
                type_key = (machine_type_code, resource_type)
                resource_capacity_by_machine_type[type_key] = max(
                    resource_capacity_by_machine_type.get(type_key, 0),
                    resource_capacity_by_machine_and_type[machine_key],
                )

    state_codes_by_machine: dict[str, set[str]] = {}
    state_features_by_case_key: dict[tuple[str, str], dict[str, str]] = {}
    for row in state_rows:
        machine_code = _required(row, "machine_code", errors)
        state_code = _required(row, "state_code", errors)
        state_type = _required(row, "state_type", errors)
        if state_type and state_type not in VALID_STATE_TYPES:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "state_type", f"Invalid state_type: {state_type}"))
        if machine_code and machine_code not in machine_to_type:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "machine_code", f"Unknown machine_code: {machine_code}"))
        features = _parse_state_features(row.data.get("features"), row, errors)
        allowed_features = state_defs_by_type.get(machine_to_type.get(machine_code, ""), set())
        for feature_key in features:
            if feature_key not in allowed_features:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "features", f"Feature '{feature_key}' is not defined for machine type"))
        state_codes_by_machine.setdefault(machine_code, set()).add(state_code)
        state_features_by_case_key[(machine_code, state_code)] = features

    rule_target_deltas: dict[str, set[str]] = {}
    for row in rule_rows:
        code = _required(row, "code", errors)
        machine_type_code = _required(row, "machine_type_code", errors)
        _required(row, "name", errors)
        duration = _parse_int(row.data.get("duration_min"))
        if duration is None or duration < 1:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "duration_min", "duration_min must be a positive integer"))
        if machine_type_code and machine_type_code not in machine_type_codes:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "machine_type_code", f"Unknown machine_type_code: {machine_type_code}"))
        activity_node_code = _text(row.data.get("activity_node_code"))
        atomic_activity_code = _text(row.data.get("atomic_activity_code"))
        if activity_node_code and atomic_activity_code:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "atomic_activity_code", "Use either activity_node_code or atomic_activity_code, not both"))
        if activity_node_code:
            activity_info = activity_node_info.get((machine_type_code, activity_node_code))
            if activity_info is None:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "activity_node_code", f"Unknown activity_node_code: {activity_node_code}"))
            elif activity_info.get("level") != 3:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "activity_node_code", "Op rules can only bind to level-3 activity nodes"))
        if atomic_activity_code and (machine_type_code, atomic_activity_code) not in atomic_activity_info:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "atomic_activity_code", f"Unknown atomic_activity_code: {atomic_activity_code}"))
        active = _parse_bool(row.data.get("is_active"), default=True)
        repair = _parse_bool(row.data.get("is_repair"), default=False)
        if active is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "is_active", "Invalid boolean value"))
        if repair is None:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "is_repair", "Invalid boolean value"))
        allowed_features = state_defs_by_type.get(machine_type_code, set())
        preconditions = _parse_preconditions(row.data.get("preconditions"), row, errors)
        effects = _parse_effects(row.data.get("effects"), row, errors)
        reqs = _parse_resource_reqs(row.data.get("resource_reqs"), row, errors)
        if not effects:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "effects", "At least one effect is required"))
        for item in preconditions:
            if item["feature_key"] not in allowed_features:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "preconditions", f"Unknown feature_key for machine type: {item['feature_key']}"))
        for item in effects:
            if item["feature_key"] not in allowed_features:
                errors.append(ImportErrorItem(row.sheet, row.row_number, "effects", f"Unknown feature_key for machine type: {item['feature_key']}"))
            else:
                rule_target_deltas.setdefault(machine_type_code, set()).add(item["feature_key"])
        for item in reqs:
            resource_type = item["resource_type"]
            if resource_type not in resource_types_by_machine_type.get(machine_type_code, set()):
                errors.append(ImportErrorItem(row.sheet, row.row_number, "resource_reqs", f"Unknown resource_type: {resource_type}"))
            elif item["quantity"] > resource_capacity_by_machine_type.get((machine_type_code, resource_type), 0):
                errors.append(ImportErrorItem(row.sheet, row.row_number, "resource_reqs", f"Resource quantity exceeds available capacity: {resource_type}"))

    solve_cases: list[dict[str, Any]] = []
    for row in solve_case_rows:
        case_code = _required(row, "case_code", errors)
        machine_code = _required(row, "machine_code", errors)
        current_state_code = _required(row, "current_state_code", errors)
        target_state_code = _required(row, "target_state_code", errors)
        objective = _text(row.data.get("objective")) or "minimize_makespan"
        if machine_code and machine_code not in machine_to_type:
            errors.append(ImportErrorItem(row.sheet, row.row_number, "machine_code", f"Unknown machine_code: {machine_code}"))
        if current_state_code and current_state_code not in state_codes_by_machine.get(machine_code, set()):
            errors.append(ImportErrorItem(row.sheet, row.row_number, "current_state_code", f"Unknown state_code for machine: {current_state_code}"))
        if target_state_code and target_state_code not in state_codes_by_machine.get(machine_code, set()):
            errors.append(ImportErrorItem(row.sheet, row.row_number, "target_state_code", f"Unknown state_code for machine: {target_state_code}"))
        for json_field in ("objectives_json", "constraints_json"):
            if not _blank(row.data.get(json_field)):
                try:
                    _parse_json(row.data.get(json_field))
                except json.JSONDecodeError:
                    errors.append(ImportErrorItem(row.sheet, row.row_number, json_field, "Invalid JSON"))
        current_features = state_features_by_case_key.get((machine_code, current_state_code), {})
        target_features = state_features_by_case_key.get((machine_code, target_state_code), {})
        deltas = {key for key, value in target_features.items() if current_features.get(key) != value}
        machine_type_code = machine_to_type.get(machine_code, "")
        if deltas and not (deltas & rule_target_deltas.get(machine_type_code, set())):
            errors.append(ImportErrorItem(row.sheet, row.row_number, "target_state_code", "No rule effect appears to address the target delta"))
        solve_cases.append(
            {
                "case_code": case_code,
                "machine_code": machine_code,
                "current_state_code": current_state_code,
                "target_state_code": target_state_code,
                "objective": objective,
            }
        )

    preview = {
        "feature_catalog": _preview(feature_rows, "feature_key", existing["features"]),
        "machine_types": _preview(machine_type_rows, "code", existing["machine_types"]),
        "machines": _preview(machine_rows, "code", existing["machines"]),
        "resources": _preview_by_machine(resource_rows, existing["resources"]),
        "activity_nodes": _preview_by_machine_type(activity_node_rows, "code", existing["activity_nodes"]),
        "atomic_activities": _preview_by_machine_type(atomic_activity_rows, "code", existing["atomic_activities"]),
        "activity_package_atomic_refs": _preview_atomic_refs(atomic_ref_rows, existing["activity_package_atomic_refs"]),
        "state_nodes": _preview_by_machine_type(state_node_rows, "code", existing["state_nodes"]),
        "scope_guards": _preview_scope_guards(scope_guard_rows, existing["scope_guards"]),
        "maintenance_intents": _preview_by_machine_type(
            maintenance_intent_rows,
            "issue_type",
            existing["maintenance_intents"],
        ),
        "layered_health_checks": {"create": len(layered_health_check_rows), "update": 0},
        "rules": _preview(rule_rows, "code", existing["rules"]),
        "state_feature_defs": await _preview_state_defs(state_def_rows, session, existing["machine_types"], machine_type_rows),
        "states": {"create": len(state_rows), "update": 0},
    }

    summary = {
        "scenario_code": scenario_code,
        "scenario_name": scenario_name,
        "feature_catalog_total": len(feature_rows),
        "machine_types_total": len(machine_type_rows),
        "machines_total": len(machine_rows),
        "state_feature_defs_total": len(state_def_rows),
        "resources_total": len(resource_rows),
        "activity_nodes_total": len(activity_node_rows),
        "atomic_activities_total": len(atomic_activity_rows),
        "activity_package_atomic_refs_total": len(atomic_ref_rows),
        "state_nodes_total": len(state_node_rows),
        "scope_guards_total": len(scope_guard_rows),
        "maintenance_intents_total": len(maintenance_intent_rows),
        "layered_health_checks_total": len(layered_health_check_rows),
        "rules_total": len(rule_rows),
        "states_total": len(state_rows),
        "solve_cases_total": len(solve_case_rows),
        "error_count": len(errors),
    }

    return {
        "summary": summary,
        "preview": preview,
        "solve_cases": solve_cases,
        "errors": [error.to_dict() for error in errors],
    }


def _preview(rows: list[ParsedRow], key_field: str, existing: dict[str, Any]) -> dict[str, int]:
    create = 0
    update = 0
    for row in rows:
        key = _text(row.data.get(key_field))
        if not key:
            continue
        if key in existing:
            update += 1
        else:
            create += 1
    return {"create": create, "update": update}


def _preview_by_machine(
    rows: list[ParsedRow],
    existing: dict[tuple[str, str], Any],
) -> dict[str, int]:
    create = 0
    update = 0
    for row in rows:
        machine_code = _text(row.data.get("machine_code"))
        code = _text(row.data.get("code"))
        if not machine_code or not code:
            continue
        if (machine_code, code) in existing:
            update += 1
        else:
            create += 1
    return {"create": create, "update": update}


def _preview_by_machine_type(
    rows: list[ParsedRow],
    code_field: str,
    existing: dict[tuple[str, str], Any],
) -> dict[str, int]:
    create = 0
    update = 0
    for row in rows:
        machine_type_code = _text(row.data.get("machine_type_code"))
        code = _text(row.data.get(code_field))
        if not machine_type_code or not code:
            continue
        if (machine_type_code, code) in existing:
            update += 1
        else:
            create += 1
    return {"create": create, "update": update}


def _preview_atomic_refs(
    rows: list[ParsedRow],
    existing: dict[tuple[str, str, str], Any],
) -> dict[str, int]:
    create = 0
    update = 0
    for row in rows:
        machine_type_code = _text(row.data.get("machine_type_code"))
        package_code = _text(row.data.get("package_code"))
        atomic_activity_code = _text(row.data.get("atomic_activity_code"))
        if not machine_type_code or not package_code or not atomic_activity_code:
            continue
        if (machine_type_code, package_code, atomic_activity_code) in existing:
            update += 1
        else:
            create += 1
    return {"create": create, "update": update}


def _preview_scope_guards(
    rows: list[ParsedRow],
    existing: dict[tuple[str, str, str], Any],
) -> dict[str, int]:
    create = 0
    update = 0
    for row in rows:
        machine_type_code = _text(row.data.get("machine_type_code"))
        activity_node_code = _text(row.data.get("activity_node_code"))
        name = _text(row.data.get("name"))
        if not machine_type_code or not activity_node_code or not name:
            continue
        if (machine_type_code, activity_node_code, name) in existing:
            update += 1
        else:
            create += 1
    return {"create": create, "update": update}


async def _preview_state_defs(
    rows: list[ParsedRow],
    session: AsyncSession,
    existing_machine_types: dict[str, MachineType],
    machine_type_rows: list[ParsedRow],
) -> dict[str, int]:
    machine_types = dict(existing_machine_types)
    for row in machine_type_rows:
        code = _text(row.data.get("code"))
        if code and code not in machine_types:
            machine_types[code] = MachineType(code=code, name=_text(row.data.get("name")))
    create = 0
    update = 0
    for row in rows:
        machine_type = machine_types.get(_text(row.data.get("machine_type_code")))
        feature_key = _text(row.data.get("feature_key"))
        if not machine_type or not feature_key or not getattr(machine_type, "id", None):
            create += 1
            continue
        result = await session.execute(
            select(StateFeatureDef.id).where(
                StateFeatureDef.machine_type_id == machine_type.id,
                StateFeatureDef.feature_key == feature_key,
            )
        )
        if result.scalar_one_or_none() is None:
            create += 1
        else:
            update += 1
    return {"create": create, "update": update}


async def import_scenario_workbook(parsed: ScenarioWorkbook, session: AsyncSession) -> dict[str, Any]:
    machine_types: dict[str, MachineType] = {}
    machines: dict[str, Machine] = {}
    activity_nodes: dict[tuple[str, str], ActivityNode] = {}
    atomic_activities: dict[tuple[str, str], AtomicActivity] = {}
    state_nodes: dict[tuple[str, str], StateNode] = {}
    states: dict[tuple[str, str], MachineState] = {}

    for row in parsed.sheet("feature_catalog"):
        feature_key = _text(row.data.get("feature_key"))
        existing = await session.get(FeatureDefinition, feature_key)
        payload = {
            "value_type": _text(row.data.get("value_type")),
            "allowed_values": _parse_allowed_values(row.data.get("allowed_values")),
            "unit": _text(row.data.get("unit")) or None,
            "description": _text(row.data.get("description")) or None,
        }
        if existing:
            existing.value_type = payload["value_type"]
            existing.allowed_values = payload["allowed_values"]
            existing.unit = payload["unit"]
            existing.description = payload["description"]
        else:
            session.add(FeatureDefinition(feature_key=feature_key, **payload))

    await session.flush()

    for row in parsed.sheet("machine_type"):
        code = _text(row.data.get("code"))
        obj = await _get_one(session, MachineType, MachineType.code == code)
        if obj is None:
            obj = MachineType(code=code, name=_text(row.data.get("name")), description=_text(row.data.get("description")) or None)
            session.add(obj)
        else:
            obj.name = _text(row.data.get("name"))
            obj.description = _text(row.data.get("description")) or None
        await session.flush()
        machine_types[code] = obj

    for row in parsed.sheet("machines"):
        code = _text(row.data.get("code"))
        machine_type = machine_types.get(_text(row.data.get("machine_type_code"))) or await _get_one(
            session, MachineType, MachineType.code == _text(row.data.get("machine_type_code"))
        )
        obj = await _get_one(session, Machine, Machine.code == code)
        if obj is None:
            obj = Machine(
                code=code,
                machine_type_id=machine_type.id,
                name=_text(row.data.get("name")),
                location=_text(row.data.get("location")) or None,
            )
            session.add(obj)
        else:
            obj.machine_type_id = machine_type.id
            obj.name = _text(row.data.get("name"))
            obj.location = _text(row.data.get("location")) or None
        await session.flush()
        machines[code] = obj

    for row in parsed.sheet("state_feature_defs"):
        machine_type = machine_types.get(_text(row.data.get("machine_type_code"))) or await _get_one(
            session, MachineType, MachineType.code == _text(row.data.get("machine_type_code"))
        )
        feature_key = _text(row.data.get("feature_key"))
        result = await session.execute(
            select(StateFeatureDef).where(
                StateFeatureDef.machine_type_id == machine_type.id,
                StateFeatureDef.feature_key == feature_key,
            )
        )
        obj = result.scalar_one_or_none()
        payload = {
            "feature_name": _text(row.data.get("feature_name")) or None,
            "value_type": _text(row.data.get("value_type")),
            "allowed_values": _parse_allowed_values(row.data.get("allowed_values")),
        }
        if obj is None:
            session.add(StateFeatureDef(machine_type_id=machine_type.id, feature_key=feature_key, **payload))
        else:
            obj.feature_name = payload["feature_name"]
            obj.value_type = payload["value_type"]
            obj.allowed_values = payload["allowed_values"]

    await session.flush()

    for row in sorted(parsed.sheet("activity_nodes"), key=lambda item: _parse_int(item.data.get("level")) or 0):
        machine_type_code = _text(row.data.get("machine_type_code"))
        code = _text(row.data.get("code"))
        machine_type = machine_types.get(machine_type_code) or await _get_one(
            session, MachineType, MachineType.code == machine_type_code
        )
        parent_code = _text(row.data.get("parent_code"))
        parent = activity_nodes.get((machine_type_code, parent_code))
        if parent is None and parent_code:
            parent = await _get_one(
                session,
                ActivityNode,
                (ActivityNode.machine_type_id == machine_type.id) & (ActivityNode.code == parent_code),
            )
        obj = await _get_one(
            session,
            ActivityNode,
            (ActivityNode.machine_type_id == machine_type.id) & (ActivityNode.code == code),
        )
        metadata = _parse_json(row.data.get("metadata_json")) if not _blank(row.data.get("metadata_json")) else None
        payload = {
            "machine_type_id": machine_type.id,
            "parent_id": parent.id if parent else None,
            "level": _parse_int(row.data.get("level")) or 1,
            "name": _text(row.data.get("name")),
            "activity_category": _text(row.data.get("activity_category")) or "normal",
            "sort_order": _parse_int(row.data.get("sort_order")) or 0,
            "is_active": bool(_parse_bool(row.data.get("is_active"), default=True)),
            "metadata_json": metadata,
        }
        if obj is None:
            obj = ActivityNode(code=code, **payload)
            session.add(obj)
            await session.flush()
        else:
            for key, value in payload.items():
                setattr(obj, key, value)
            await session.flush()
        activity_nodes[(machine_type_code, code)] = obj

    for row in parsed.sheet("atomic_activities"):
        machine_type_code = _text(row.data.get("machine_type_code"))
        code = _text(row.data.get("code"))
        machine_type = machine_types.get(machine_type_code) or await _get_one(
            session, MachineType, MachineType.code == machine_type_code
        )
        obj = await _get_one(
            session,
            AtomicActivity,
            (AtomicActivity.machine_type_id == machine_type.id) & (AtomicActivity.code == code),
        )
        metadata = _parse_json(row.data.get("metadata_json")) if not _blank(row.data.get("metadata_json")) else None
        payload = {
            "machine_type_id": machine_type.id,
            "name": _text(row.data.get("name")),
            "activity_category": _text(row.data.get("activity_category")) or "normal",
            "sort_order": _parse_int(row.data.get("sort_order")) or 0,
            "is_active": bool(_parse_bool(row.data.get("is_active"), default=True)),
            "metadata_json": metadata,
        }
        if obj is None:
            obj = AtomicActivity(code=code, **payload)
            session.add(obj)
            await session.flush()
        else:
            for key, value in payload.items():
                setattr(obj, key, value)
            await session.flush()
        atomic_activities[(machine_type_code, code)] = obj

    for (machine_type_code, code), node in list(activity_nodes.items()):
        if node.level != 3:
            continue
        machine_type = machine_types.get(machine_type_code) or await _get_one(
            session, MachineType, MachineType.code == machine_type_code
        )
        atomic = atomic_activities.get((machine_type_code, code))
        if atomic is None:
            atomic = await _get_one(
                session,
                AtomicActivity,
                (AtomicActivity.machine_type_id == machine_type.id) & (AtomicActivity.code == code),
            )
        if atomic is None:
            atomic = AtomicActivity(
                machine_type_id=machine_type.id,
                code=code,
                name=node.name,
                activity_category=node.activity_category,
                sort_order=node.sort_order,
                is_active=node.is_active,
                metadata_json=node.metadata_json,
            )
            session.add(atomic)
            await session.flush()
        atomic_activities[(machine_type_code, code)] = atomic
        if node.parent_id:
            ref = await _get_one(
                session,
                ActivityPackageAtomicRef,
                (ActivityPackageAtomicRef.activity_node_id == node.parent_id)
                & (ActivityPackageAtomicRef.atomic_activity_id == atomic.id),
            )
            if ref is None:
                session.add(
                    ActivityPackageAtomicRef(
                        activity_node_id=node.parent_id,
                        atomic_activity_id=atomic.id,
                        sort_order=node.sort_order,
                        is_active=node.is_active,
                        metadata_json=None,
                    )
                )

    await session.flush()

    for row in parsed.sheet("activity_package_atomic_refs"):
        machine_type_code = _text(row.data.get("machine_type_code"))
        package_code = _text(row.data.get("package_code"))
        atomic_activity_code = _text(row.data.get("atomic_activity_code"))
        package = activity_nodes.get((machine_type_code, package_code))
        if package is None:
            machine_type = machine_types.get(machine_type_code) or await _get_one(
                session, MachineType, MachineType.code == machine_type_code
            )
            package = await _get_one(
                session,
                ActivityNode,
                (ActivityNode.machine_type_id == machine_type.id) & (ActivityNode.code == package_code),
            )
        atomic = atomic_activities.get((machine_type_code, atomic_activity_code))
        if atomic is None:
            machine_type = machine_types.get(machine_type_code) or await _get_one(
                session, MachineType, MachineType.code == machine_type_code
            )
            atomic = await _get_one(
                session,
                AtomicActivity,
                (AtomicActivity.machine_type_id == machine_type.id) & (AtomicActivity.code == atomic_activity_code),
            )
        ref = await _get_one(
            session,
            ActivityPackageAtomicRef,
            (ActivityPackageAtomicRef.activity_node_id == package.id)
            & (ActivityPackageAtomicRef.atomic_activity_id == atomic.id),
        )
        metadata = _parse_json(row.data.get("metadata_json")) if not _blank(row.data.get("metadata_json")) else None
        payload = {
            "activity_node_id": package.id,
            "atomic_activity_id": atomic.id,
            "sort_order": _parse_int(row.data.get("sort_order")) or 0,
            "is_active": bool(_parse_bool(row.data.get("is_active"), default=True)),
            "metadata_json": metadata,
        }
        if ref is None:
            session.add(ActivityPackageAtomicRef(**payload))
        else:
            for key, value in payload.items():
                setattr(ref, key, value)

    await session.flush()

    for row in sorted(parsed.sheet("state_nodes"), key=lambda item: _parse_int(item.data.get("level")) or 0):
        machine_type_code = _text(row.data.get("machine_type_code"))
        code = _text(row.data.get("code"))
        machine_type = machine_types.get(machine_type_code) or await _get_one(
            session, MachineType, MachineType.code == machine_type_code
        )
        parent_code = _text(row.data.get("parent_code"))
        parent = state_nodes.get((machine_type_code, parent_code))
        if parent is None and parent_code:
            parent = await _get_one(
                session,
                StateNode,
                (StateNode.machine_type_id == machine_type.id) & (StateNode.code == parent_code),
            )
        level = _parse_int(row.data.get("level")) or 1
        obj = await _get_one(
            session,
            StateNode,
            (StateNode.machine_type_id == machine_type.id) & (StateNode.code == code),
        )
        metadata = _parse_json(row.data.get("metadata_json")) if not _blank(row.data.get("metadata_json")) else None
        feature_key = _text(row.data.get("feature_key")) or None
        target_value = _text(row.data.get("target_value")) or None
        state_kind = _text(row.data.get("state_kind")) or ("aggregate" if not feature_key and not target_value else "atomic")
        if state_kind != "aggregate" and feature_key and target_value:
            await _ensure_import_state_feature_def(
                session,
                machine_type,
                feature_key=feature_key,
                feature_name=_text(row.data.get("name")),
                target_value=target_value,
            )
        payload = {
            "machine_type_id": machine_type.id,
            "parent_id": parent.id if parent else None,
            "level": level,
            "name": _text(row.data.get("name")),
            "feature_key": feature_key,
            "operator": _text(row.data.get("operator")) or "eq",
            "target_value": target_value,
            "state_kind": state_kind,
            "sort_order": _parse_int(row.data.get("sort_order")) or 0,
            "is_active": bool(_parse_bool(row.data.get("is_active"), default=True)),
            "metadata_json": metadata,
        }
        if obj is None:
            obj = StateNode(code=code, **payload)
            session.add(obj)
            await session.flush()
        else:
            for key, value in payload.items():
                setattr(obj, key, value)
            await session.flush()
        state_nodes[(machine_type_code, code)] = obj

    for row in parsed.sheet("resources"):
        machine_code = _text(row.data.get("machine_code"))
        code = _text(row.data.get("code"))
        machine = machines.get(machine_code) or await _get_one(session, Machine, Machine.code == machine_code)
        obj = await _get_one(
            session,
            Resource,
            (Resource.machine_id == machine.id) & (Resource.code == code),
        )
        meta = _parse_json(row.data.get("meta_json")) if not _blank(row.data.get("meta_json")) else None
        payload = {
            "machine_id": machine.id,
            "name": _text(row.data.get("name")),
            "resource_type": _text(row.data.get("resource_type")),
            "capacity": _parse_int(row.data.get("capacity")) or 1,
            "is_available": _parse_bool(row.data.get("is_available"), default=True),
            "meta": meta,
        }
        if obj is None:
            session.add(Resource(code=code, **payload))
        else:
            obj.machine_id = payload["machine_id"]
            obj.name = payload["name"]
            obj.resource_type = payload["resource_type"]
            obj.capacity = payload["capacity"]
            obj.is_available = bool(payload["is_available"])
            obj.meta = payload["meta"]

    await session.flush()

    for row in parsed.sheet("scope_guards"):
        machine_type_code = _text(row.data.get("machine_type_code"))
        activity_node_code = _text(row.data.get("activity_node_code"))
        machine_type = machine_types.get(machine_type_code) or await _get_one(
            session, MachineType, MachineType.code == machine_type_code
        )
        activity_node = activity_nodes.get((machine_type_code, activity_node_code))
        if activity_node is None:
            activity_node = await _get_one(
                session,
                ActivityNode,
                (ActivityNode.machine_type_id == machine_type.id) & (ActivityNode.code == activity_node_code),
            )
        name = _text(row.data.get("name"))
        guard = await _get_one(
            session,
            ScopeGuard,
            (ScopeGuard.activity_node_id == activity_node.id) & (ScopeGuard.name == name),
        )
        metadata = _parse_json(row.data.get("metadata_json")) if not _blank(row.data.get("metadata_json")) else None
        payload = {
            "activity_node_id": activity_node.id,
            "name": name,
            "description": _text(row.data.get("description")) or None,
            "is_active": bool(_parse_bool(row.data.get("is_active"), default=True)),
            "metadata_json": metadata,
        }
        if guard is None:
            guard = ScopeGuard(**payload)
            session.add(guard)
            await session.flush()
        else:
            for key, value in payload.items():
                setattr(guard, key, value)
            await session.flush()
        await session.execute(delete(ScopeGuardPrecond).where(ScopeGuardPrecond.scope_guard_id == guard.id))
        preconditions = _parse_scope_guard_preconditions(row.data.get("preconditions"), row, [])
        for item in preconditions:
            state_node_code = item["state_node_code"]
            state_node = state_nodes.get((machine_type_code, state_node_code))
            if state_node is None:
                state_node = await _get_one(
                    session,
                    StateNode,
                    (StateNode.machine_type_id == machine_type.id) & (StateNode.code == state_node_code),
                )
            session.add(
                ScopeGuardPrecond(
                    scope_guard_id=guard.id,
                    state_node_id=state_node.id,
                    operator=item["operator"],
                    expected_value=item["expected_value"],
                    value_list=item["value_list"],
                )
            )

    await session.flush()

    maintenance_intent_templates: list[dict[str, Any]] = []
    for row in parsed.sheet("maintenance_intents"):
        machine_type_code = _text(row.data.get("machine_type_code"))
        machine_type = machine_types.get(machine_type_code) or await _get_one(
            session, MachineType, MachineType.code == machine_type_code
        )
        scope_activity_node_code = _text(row.data.get("scope_activity_node_code"))
        scope_activity_node = activity_nodes.get((machine_type_code, scope_activity_node_code))
        if scope_activity_node is None:
            scope_activity_node = await _get_one(
                session,
                ActivityNode,
                (ActivityNode.machine_type_id == machine_type.id)
                & (ActivityNode.code == scope_activity_node_code),
            )

        target_state_node_ids: list[int] = []
        for state_node_code in _parse_code_list(row.data.get("target_state_node_codes")):
            state_node = state_nodes.get((machine_type_code, state_node_code))
            if state_node is None:
                state_node = await _get_one(
                    session,
                    StateNode,
                    (StateNode.machine_type_id == machine_type.id) & (StateNode.code == state_node_code),
                )
            target_state_node_ids.append(state_node.id)

        candidate_scope_codes = _parse_code_list(row.data.get("candidate_activity_scope_codes"))
        candidate_activity_scope_ids: list[int] = []
        for activity_node_code in candidate_scope_codes:
            activity_node = activity_nodes.get((machine_type_code, activity_node_code))
            if activity_node is None:
                activity_node = await _get_one(
                    session,
                    ActivityNode,
                    (ActivityNode.machine_type_id == machine_type.id) & (ActivityNode.code == activity_node_code),
                )
            candidate_activity_scope_ids.append(activity_node.id)
        if not candidate_activity_scope_ids:
            candidate_activity_scope_ids = [scope_activity_node.id]

        issue_type = _text(row.data.get("issue_type"))
        obj = await _get_one(
            session,
            MaintenanceIntentTemplate,
            (MaintenanceIntentTemplate.machine_type_id == machine_type.id)
            & (MaintenanceIntentTemplate.issue_type == issue_type),
        )
        metadata = _parse_json(row.data.get("metadata_json")) if not _blank(row.data.get("metadata_json")) else None
        observed_facts = _parse_maintenance_fact_templates(
            row.data.get("observed_fact_templates"),
            row,
            "observed_fact_templates",
            [],
        )
        desired_facts = _parse_maintenance_fact_templates(
            row.data.get("desired_fact_templates"),
            row,
            "desired_fact_templates",
            [],
        )
        payload = {
            "machine_type_id": machine_type.id,
            "scope_activity_node_id": scope_activity_node.id,
            "issue_type": issue_type,
            "name": _text(row.data.get("name")),
            "description": _text(row.data.get("description")) or None,
            "target_state_node_ids": list(dict.fromkeys(target_state_node_ids)),
            "candidate_activity_scope_ids": list(dict.fromkeys(candidate_activity_scope_ids)),
            "observed_fact_templates": observed_facts,
            "desired_fact_templates": desired_facts,
            "is_active": bool(_parse_bool(row.data.get("is_active"), default=True)),
            "metadata_json": metadata,
        }
        if obj is None:
            obj = MaintenanceIntentTemplate(**payload)
            session.add(obj)
            await session.flush()
        else:
            for key, value in payload.items():
                setattr(obj, key, value)
            await session.flush()
        maintenance_intent_templates.append(
            {
                "issue_type": obj.issue_type,
                "name": obj.name,
                "id": obj.id,
                "machine_type_code": machine_type_code,
                "scope_activity_node_code": scope_activity_node_code,
            }
        )

    await session.flush()

    for row in parsed.sheet("rules"):
        code = _text(row.data.get("code"))
        machine_type_code = _text(row.data.get("machine_type_code"))
        machine_type = machine_types.get(machine_type_code) or await _get_one(
            session, MachineType, MachineType.code == machine_type_code
        )
        activity_node_code = _text(row.data.get("activity_node_code"))
        atomic_activity_code = _text(row.data.get("atomic_activity_code"))
        activity_node = activity_nodes.get((machine_type_code, activity_node_code))
        if activity_node is None and activity_node_code:
            activity_node = await _get_one(
                session,
                ActivityNode,
                (ActivityNode.machine_type_id == machine_type.id) & (ActivityNode.code == activity_node_code),
            )
        atomic_activity = None
        if atomic_activity_code:
            atomic_activity = atomic_activities.get((machine_type_code, atomic_activity_code))
            if atomic_activity is None:
                atomic_activity = await _get_one(
                    session,
                    AtomicActivity,
                    (AtomicActivity.machine_type_id == machine_type.id) & (AtomicActivity.code == atomic_activity_code),
                )
        elif activity_node is not None and activity_node.level == 3:
            atomic_activity = atomic_activities.get((machine_type_code, activity_node.code))
            if atomic_activity is None:
                atomic_activity = await _get_one(
                    session,
                    AtomicActivity,
                    (AtomicActivity.machine_type_id == machine_type.id) & (AtomicActivity.code == activity_node.code),
                )
        rule = await _get_one(session, OpRule, OpRule.code == code)
        payload = {
            "machine_type_id": machine_type.id,
            "activity_node_id": activity_node.id if activity_node and not atomic_activity_code else None,
            "atomic_activity_id": atomic_activity.id if atomic_activity else None,
            "name": _text(row.data.get("name")),
            "duration_min": _parse_int(row.data.get("duration_min")) or 1,
            "description": _text(row.data.get("description")) or None,
            "is_active": bool(_parse_bool(row.data.get("is_active"), default=True)),
            "is_repair": bool(_parse_bool(row.data.get("is_repair"), default=False)),
        }
        if rule is None:
            rule = OpRule(code=code, **payload)
            session.add(rule)
            await session.flush()
        else:
            for key, value in payload.items():
                setattr(rule, key, value)
            await session.flush()
        await _replace_rule_children(rule, row, session)

    await session.flush()

    post_import_health_checks = await _run_post_import_health_checks(
        parsed,
        session,
        machine_types,
        activity_nodes,
        state_nodes,
    )

    for row in parsed.sheet("states"):
        machine = machines.get(_text(row.data.get("machine_code"))) or await _get_one(
            session, Machine, Machine.code == _text(row.data.get("machine_code"))
        )
        state = MachineState(
            machine_id=machine.id,
            state_type=_text(row.data.get("state_type")),
            label=_text(row.data.get("label")) or _text(row.data.get("state_code")),
        )
        session.add(state)
        await session.flush()
        features = _parse_state_features(row.data.get("features"), row, [])
        session.add_all(
            [
                MachineStateFeature(machine_state_id=state.id, feature_key=feature_key, feature_value=feature_value)
                for feature_key, feature_value in features.items()
            ]
        )
        states[(_text(row.data.get("machine_code")), _text(row.data.get("state_code")))] = state

    await session.flush()

    solve_cases: list[dict[str, Any]] = []
    for row in parsed.sheet("solve_cases"):
        machine_code = _text(row.data.get("machine_code"))
        current_state_code = _text(row.data.get("current_state_code"))
        target_state_code = _text(row.data.get("target_state_code"))
        machine = machines.get(machine_code) or await _get_one(session, Machine, Machine.code == machine_code)
        current_state = states.get((machine_code, current_state_code))
        target_state = states.get((machine_code, target_state_code))
        solve_cases.append(
            {
                "case_code": _text(row.data.get("case_code")),
                "machine_code": machine_code,
                "machine_id": machine.id if machine else None,
                "current_state_code": current_state_code,
                "current_state_id": current_state.id if current_state else None,
                "target_state_code": target_state_code,
                "target_state_id": target_state.id if target_state else None,
                "objective": _text(row.data.get("objective")) or "minimize_makespan",
            }
        )

    return {
        "solve_cases": solve_cases,
        "maintenance_intent_templates": maintenance_intent_templates,
        "post_import_health_checks": post_import_health_checks,
    }


async def _get_one(session: AsyncSession, model: type, condition: Any) -> Any:
    result = await session.execute(select(model).where(condition).limit(1))
    return result.scalar_one_or_none()


def _append_allowed_value(allowed_values: Any, value: str) -> list[Any]:
    if allowed_values is None:
        return [value]
    if isinstance(allowed_values, list):
        values = list(allowed_values)
    elif isinstance(allowed_values, dict):
        values = list(allowed_values.get("values", allowed_values.values()))
    else:
        values = [allowed_values]
    if value not in {str(item) for item in values}:
        values.append(value)
    return values


async def _ensure_import_state_feature_def(
    session: AsyncSession,
    machine_type: MachineType,
    *,
    feature_key: str,
    feature_name: str,
    target_value: str,
) -> None:
    global_def = await session.get(FeatureDefinition, feature_key)
    if global_def is None:
        session.add(
            FeatureDefinition(
                feature_key=feature_key,
                value_type="enum",
                allowed_values=[target_value],
                description=f"Auto-created from imported state leaf '{feature_name}'",
            )
        )
    elif global_def.value_type == "enum":
        global_def.allowed_values = _append_allowed_value(global_def.allowed_values, target_value)

    state_def = await _get_one(
        session,
        StateFeatureDef,
        (StateFeatureDef.machine_type_id == machine_type.id)
        & (StateFeatureDef.feature_key == feature_key),
    )
    if state_def is None:
        session.add(
            StateFeatureDef(
                machine_type_id=machine_type.id,
                feature_key=feature_key,
                feature_name=feature_name,
                value_type="enum",
                allowed_values=[target_value],
            )
        )
    elif state_def.value_type == "enum":
        state_def.allowed_values = _append_allowed_value(state_def.allowed_values, target_value)


def _compact_health_result(health: dict[str, Any]) -> dict[str, Any]:
    diagnostics = health.get("diagnostics") or []
    blocking = [item for item in diagnostics if item.get("severity") == "error"]
    warnings = [item for item in diagnostics if item.get("severity") != "error"]
    return {
        "status": health.get("status", "unknown"),
        "summary": health.get("summary", {}),
        "blocking_count": len(blocking),
        "warning_count": len(warnings),
        "diagnostics": [
            {
                "code": item.get("code"),
                "severity": item.get("severity", "warning"),
                "message": item.get("message"),
                "feature_key": item.get("feature_key"),
                "operator": item.get("operator"),
                "target_value": item.get("target_value"),
                "provider_count": item.get("provider_count"),
                "op_rule_id": item.get("op_rule_id"),
                "activity_node_id": item.get("activity_node_id"),
                "state_node_id": item.get("state_node_id"),
                "source_type": item.get("source_type"),
            }
            for item in diagnostics
        ],
    }


async def _run_post_import_health_checks(
    parsed: ScenarioWorkbook,
    session: AsyncSession,
    machine_types: dict[str, MachineType],
    activity_nodes: dict[tuple[str, str], ActivityNode],
    state_nodes: dict[tuple[str, str], StateNode],
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for row in parsed.sheet("layered_health_checks"):
        machine_type_code = _text(row.data.get("machine_type_code"))
        machine_type = machine_types.get(machine_type_code) or await _get_one(
            session,
            MachineType,
            MachineType.code == machine_type_code,
        )
        target_state_node_ids: list[int] = []
        for state_node_code in _parse_code_list(row.data.get("target_state_node_codes")):
            state_node = state_nodes.get((machine_type_code, state_node_code))
            if state_node is None:
                state_node = await _get_one(
                    session,
                    StateNode,
                    (StateNode.machine_type_id == machine_type.id) & (StateNode.code == state_node_code),
                )
            target_state_node_ids.append(state_node.id)

        activity_scope_node_ids: list[int] = []
        for activity_node_code in _parse_code_list(row.data.get("activity_scope_node_codes")):
            activity_node = activity_nodes.get((machine_type_code, activity_node_code))
            if activity_node is None:
                activity_node = await _get_one(
                    session,
                    ActivityNode,
                    (ActivityNode.machine_type_id == machine_type.id) & (ActivityNode.code == activity_node_code),
                )
            activity_scope_node_ids.append(activity_node.id)

        payload = LayeredExpansionRequest(
            target_state_node_ids=list(dict.fromkeys(target_state_node_ids)),
            activity_scope_node_ids=list(dict.fromkeys(activity_scope_node_ids)),
            include_inactive=bool(_parse_bool(row.data.get("include_inactive"), default=False)),
        )
        health = await check_layered_health(session, machine_type.id, payload)
        results.append(
            {
                "check_code": _text(row.data.get("check_code")),
                "name": _text(row.data.get("name")),
                "description": _text(row.data.get("description")) or None,
                "machine_type_code": machine_type_code,
                "machine_type_id": machine_type.id,
                "target_state_node_ids": payload.target_state_node_ids,
                "activity_scope_node_ids": payload.activity_scope_node_ids,
                "include_inactive": payload.include_inactive,
                **_compact_health_result(health),
            }
        )
    return results


async def _replace_rule_children(rule: OpRule, row: ParsedRow, session: AsyncSession) -> None:
    await session.execute(delete(OpRulePrecond).where(OpRulePrecond.op_rule_id == rule.id))
    await session.execute(delete(OpRuleEffect).where(OpRuleEffect.op_rule_id == rule.id))
    await session.execute(delete(OpRuleResourceReq).where(OpRuleResourceReq.op_rule_id == rule.id))
    errors: list[ImportErrorItem] = []
    for item in _parse_preconditions(row.data.get("preconditions"), row, errors):
        session.add(
            OpRulePrecond(
                op_rule_id=rule.id,
                feature_key=item["feature_key"],
                operator=item["operator"],
                feature_value=item["feature_value"],
                value_list=item["value_list"],
            )
        )
    for item in _parse_effects(row.data.get("effects"), row, errors):
        session.add(
            OpRuleEffect(
                op_rule_id=rule.id,
                feature_key=item["feature_key"],
                effect_type=item["effect_type"],
                new_value=item["new_value"],
                delta_value=item["delta_value"],
            )
        )
    for item in _parse_resource_reqs(row.data.get("resource_reqs"), row, errors):
        session.add(
            OpRuleResourceReq(
                op_rule_id=rule.id,
                resource_type=item["resource_type"],
                quantity=item["quantity"],
                is_required=item["is_required"],
            )
        )


def build_scenario_template() -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name in REQUIRED_SHEETS + OPTIONAL_SHEETS:
        worksheet = workbook.create_sheet(sheet_name)
        if sheet_name in SHEET_FIELDS:
            fields = [*SHEET_FIELDS[sheet_name]]
            if sheet_name == "rules":
                fields.append("activity_node_code")
                fields.append("atomic_activity_code")
            worksheet.append(fields)
        elif sheet_name == "instructions":
            worksheet.append(["section", "description"])
            worksheet.append([
                "rules.effects.effect_type",
                "Allowed effect_type values: set, increment, decrement, sub, reset. For increment/decrement/sub, value is numeric delta. For set/reset, value is the target value.",
            ])
            worksheet.append(["填写流程", "先填 meta，再填特征、设备、资源、规则、状态和 solve_cases。"])
            worksheet.append(["rules.preconditions", "格式：feature_key:operator:value，多项用分号分隔。示例：prep_done:eq:true"])
            worksheet.append(["rules.effects", "格式：feature_key:effect_type:value。示例：wing_joined:set:true"])
            worksheet.append(["resources.machine_code", "必填：每个资源必须绑定一台具体机器，资源不再是全局资源。"])
            worksheet.append(["rules.resource_reqs", "格式：resource_type:quantity:is_required。示例：technician:2:true"])
            worksheet.append(["rules.activity_node_code", "Optional legacy binding to a level-3 activity_node; import will also backfill atomic_activity_id."])
            worksheet.append(["rules.atomic_activity_code", "Optional preferred binding to atomic_activities.code. Do not use together with activity_node_code."])
            worksheet.append(["activity_nodes", "Optional package tree: level 1/2 packages; legacy level-3 rows are auto-converted to atomic activities and refs."])
            worksheet.append(["atomic_activities", "Optional reusable executable activity library."])
            worksheet.append(["activity_package_atomic_refs", "Optional refs from level-2 package_code to atomic_activity_code."])
            worksheet.append(["state_nodes", "Optional arbitrary-depth state tree; aggregate rows leave fact fields empty, leaf rows use feature_key/operator=eq/target_value."])
            worksheet.append(["scope_guards.preconditions", "可选：格式 state_node_code:operator[:expected]，示例 PREP_DONE:completed。"])
            worksheet.append(["states.features", "格式：feature_key:value，多项用分号分隔。"])
            worksheet.append(["maintenance_intents", "Optional: level-2 scope_activity_node_code; comma/semicolon target_state_node_codes and candidate_activity_scope_codes; facts as feature:eq:value or JSON array."])
            worksheet.append(["layered_health_checks", "Optional post-import diagnostics: target_state_node_codes and activity_scope_node_codes use comma or semicolon separated node codes."])
        elif sheet_name == "rule_groups":
            worksheet.append(["group_code", "group_name", "rule_codes", "description"])
        elif sheet_name == "notes":
            worksheet.append(["note"])

    workbook["meta"].append(["SCENARIO_001", "业务端到端测试场景", "v1", "scenario_upsert"])
    workbook["feature_catalog"].append(["prep_done", "enum", "false,true", "", "准备完成"])
    workbook["feature_catalog"].append(["delivery_ready", "enum", "false,true", "", "交付就绪"])
    workbook["machine_type"].append(["BUSINESS_OBJECT", "业务对象", "端到端测试对象"])
    workbook["machines"].append(["BO-001", "BUSINESS_OBJECT", "业务对象 001", "测试线"])
    workbook["state_feature_defs"].append(["BUSINESS_OBJECT", "prep_done", "准备完成", "enum", "false,true"])
    workbook["state_feature_defs"].append(["BUSINESS_OBJECT", "delivery_ready", "交付就绪", "enum", "false,true"])
    workbook["resources"].append(["BO-001", "TECH-01", "技术员 01", "technician", 1, "true", ""])
    workbook["activity_nodes"].append(["BUSINESS_OBJECT", "BUSINESS_FLOW", "", 1, "业务流程", "normal", 0, "true", ""])
    workbook["activity_nodes"].append(["BUSINESS_OBJECT", "PREP_PACK", "BUSINESS_FLOW", 2, "准备包", "normal", 10, "true", ""])
    workbook["activity_nodes"].append(["BUSINESS_OBJECT", "DELIVERY_PACK", "BUSINESS_FLOW", 2, "交付包", "normal", 20, "true", ""])
    workbook["atomic_activities"].append(["BUSINESS_OBJECT", "PREP_STEP", "准备步骤", "normal", 10, "true", ""])
    workbook["atomic_activities"].append(["BUSINESS_OBJECT", "DELIVER_STEP", "交付步骤", "normal", 20, "true", ""])
    workbook["activity_package_atomic_refs"].append(["BUSINESS_OBJECT", "PREP_PACK", "PREP_STEP", 10, "true", ""])
    workbook["activity_package_atomic_refs"].append(["BUSINESS_OBJECT", "DELIVERY_PACK", "DELIVER_STEP", 20, "true", ""])
    workbook["state_nodes"].append(["BUSINESS_OBJECT", "BUSINESS_DONE", "", 1, "业务完成", "", "", "", "aggregate", 0, "true", ""])
    workbook["state_nodes"].append(["BUSINESS_OBJECT", "BUSINESS_READY", "BUSINESS_DONE", 2, "业务就绪", "", "", "", "aggregate", 0, "true", ""])
    workbook["state_nodes"].append(["BUSINESS_OBJECT", "PREP_DONE", "BUSINESS_READY", 3, "准备完成", "prep_done", "eq", "true", "atomic", 10, "true", ""])
    workbook["state_nodes"].append(["BUSINESS_OBJECT", "DELIVERY_READY", "BUSINESS_READY", 3, "交付就绪", "delivery_ready", "eq", "true", "atomic", 20, "true", ""])
    workbook["scope_guards"].append(["BUSINESS_OBJECT", "DELIVERY_PACK", "交付前准备", "", "true", "PREP_DONE:completed", ""])
    workbook["rules"].append(["OP_PREP", "BUSINESS_OBJECT", "准备", 30, "", "true", "false", "prep_done:eq:false", "prep_done:set:true", "technician:1:true", "", "PREP_STEP"])
    workbook["rules"].append(["OP_DELIVER", "BUSINESS_OBJECT", "交付", 20, "", "true", "false", "prep_done:eq:true", "delivery_ready:set:true", "technician:1:true", "", "DELIVER_STEP"])
    workbook["states"].append(["BO-001", "START", "current", "起点", "prep_done:false;delivery_ready:false"])
    workbook["states"].append(["BO-001", "TARGET", "target", "目标", "prep_done:true;delivery_ready:true"])
    workbook["solve_cases"].append(["FULL_FLOW", "BO-001", "START", "TARGET", "minimize_makespan", "", "", 2, 60])

    workbook["maintenance_intents"].append([
        "BUSINESS_OBJECT",
        "DELIVERY_MAINTENANCE",
        "Delivery maintenance",
        "DELIVERY_PACK",
        "",
        "DELIVERY_READY",
        "PREP_PACK;DELIVERY_PACK",
        "prep_done:eq:false",
        "",
        "true",
        "",
    ])
    workbook["layered_health_checks"].append([
        "BUSINESS_OBJECT",
        "DELIVERY_READY_CHECK",
        "Delivery ready health",
        "DELIVERY_READY",
        "PREP_PACK;DELIVERY_PACK",
        "false",
        "Run after import to verify target providers and Scope Guard chains.",
    ])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
