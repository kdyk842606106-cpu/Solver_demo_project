from __future__ import annotations

import asyncio
import json
import sys
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select


PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.db.models import (
    ActivityNode,
    Machine,
    MachineState,
    MachineStateFeature,
    MaintenanceIntentTemplate,
    StateNode,
)
from app.db.schemas import LayeredSolveRequest, MaintenanceSolveRequest
from app.db.session import AsyncSessionLocal
from app.services.layered_solve import solve_layered
from app.services.maintenance_solve import solve_maintenance
from app.services.scenario_import import (
    build_scenario_template,
    import_scenario_workbook,
    parse_scenario_workbook,
    validate_scenario_workbook,
)


OUT_DIR = PROJECT_ROOT / "outputs" / "layered_activity_state_demo"
XLSX_PATH = OUT_DIR / "layered_activity_state_demo.xlsx"
SUMMARY_PATH = OUT_DIR / "layered_activity_state_demo_result.json"

MACHINE_TYPE_CODE = "LAYERED_MAINT_DEMO"
MACHINE_CODE = "LAYERED-MAINT-001"


def _append_rows(sheet: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)


def build_workbook() -> bytes:
    workbook = load_workbook(BytesIO(build_scenario_template()))
    for worksheet in workbook.worksheets:
        if worksheet.max_row > 1:
            worksheet.delete_rows(2, worksheet.max_row - 1)

    _append_rows(
        workbook["meta"],
        [[
            "LAYERED_ACTIVITY_STATE_DEMO",
            "Layered Activity State Maintenance Demo",
            "v1",
            "scenario_upsert",
        ]],
    )

    _append_rows(
        workbook["feature_catalog"],
        [
            ["safe", "enum", "no,yes", "", "Shared safety isolation fact"],
            ["pump_diagnosed", "enum", "no,yes", "", "Pump diagnosis completed"],
            ["pump_fixed", "enum", "no,yes", "", "Pump repair completed"],
            ["valve_diagnosed", "enum", "no,yes", "", "Valve diagnosis completed"],
            ["valve_fixed", "enum", "no,yes", "", "Valve repair completed"],
            ["final_test", "enum", "not_run,passed", "", "Joint acceptance test result"],
        ],
    )

    _append_rows(
        workbook["machine_type"],
        [[MACHINE_TYPE_CODE, "Layered Maintenance Demo Machine", "Fixture for layered activity/state flows"]],
    )
    _append_rows(
        workbook["machines"],
        [[MACHINE_CODE, MACHINE_TYPE_CODE, "Layered Maintenance Demo 001", "Demo Line A"]],
    )

    _append_rows(
        workbook["state_feature_defs"],
        [
            [MACHINE_TYPE_CODE, "safe", "Safe isolation", "enum", "no,yes"],
            [MACHINE_TYPE_CODE, "pump_diagnosed", "Pump diagnosed", "enum", "no,yes"],
            [MACHINE_TYPE_CODE, "pump_fixed", "Pump fixed", "enum", "no,yes"],
            [MACHINE_TYPE_CODE, "valve_diagnosed", "Valve diagnosed", "enum", "no,yes"],
            [MACHINE_TYPE_CODE, "valve_fixed", "Valve fixed", "enum", "no,yes"],
            [MACHINE_TYPE_CODE, "final_test", "Final test", "enum", "not_run,passed"],
        ],
    )

    _append_rows(
        workbook["resources"],
        [
            ["LMD_TECH_A", "Layered demo technician A", "TECHNICIAN", 2, "true", ""],
        ],
    )

    _append_rows(
        workbook["activity_nodes"],
        [
            [MACHINE_TYPE_CODE, "LMD_SYSTEM", "", 1, "分层维修系统", "normal", 0, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_PREP", "LMD_SYSTEM", 2, "公共安全准备", "normal", 10, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_PUMP_MAINT", "LMD_SYSTEM", 2, "泵维修维护", "maintenance", 20, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_VALVE_MAINT", "LMD_SYSTEM", 2, "阀维修维护", "maintenance", 30, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_ACCEPTANCE", "LMD_SYSTEM", 2, "联合验收", "normal", 40, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_ISOLATE", "LMD_PREP", 3, "安全隔离", "normal", 10, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_PUMP_DIAGNOSE", "LMD_PUMP_MAINT", 3, "泵故障诊断", "maintenance", 10, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_PUMP_REPAIR", "LMD_PUMP_MAINT", 3, "泵修复", "maintenance", 20, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_VALVE_DIAGNOSE", "LMD_VALVE_MAINT", 3, "阀故障诊断", "maintenance", 10, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_VALVE_REPAIR", "LMD_VALVE_MAINT", 3, "阀修复", "maintenance", 20, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_FINAL_TEST", "LMD_ACCEPTANCE", 3, "联合功能测试", "normal", 10, "true", ""],
        ],
    )

    _append_rows(
        workbook["state_nodes"],
        [
            [MACHINE_TYPE_CODE, "LMD_OPERATION_READY", "", 1, "设备恢复可运行", "", "", "", "aggregate", 0, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_SAFETY_READY", "LMD_OPERATION_READY", 2, "安全准备完成", "", "", "", "aggregate", 10, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_REPAIR_DONE", "LMD_OPERATION_READY", 2, "维修完成", "", "", "", "aggregate", 20, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_ACCEPT_DONE", "LMD_OPERATION_READY", 2, "验收完成", "", "", "", "aggregate", 30, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_SAFE_YES", "LMD_SAFETY_READY", 3, "已安全隔离", "safe", "eq", "yes", "atomic", 10, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_PUMP_DIAG_YES", "LMD_REPAIR_DONE", 3, "泵已诊断", "pump_diagnosed", "eq", "yes", "atomic", 20, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_PUMP_FIXED_YES", "LMD_REPAIR_DONE", 3, "泵已修复", "pump_fixed", "eq", "yes", "atomic", 30, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_VALVE_DIAG_YES", "LMD_REPAIR_DONE", 3, "阀已诊断", "valve_diagnosed", "eq", "yes", "atomic", 40, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_VALVE_FIXED_YES", "LMD_REPAIR_DONE", 3, "阀已修复", "valve_fixed", "eq", "yes", "atomic", 50, "true", ""],
            [MACHINE_TYPE_CODE, "LMD_FINAL_TEST_PASSED", "LMD_ACCEPT_DONE", 3, "联合测试通过", "final_test", "eq", "passed", "atomic", 60, "true", ""],
        ],
    )

    _append_rows(
        workbook["scope_guards"],
        [
            [MACHINE_TYPE_CODE, "LMD_PUMP_MAINT", "泵维修前需要安全隔离", "", "true", "LMD_SAFE_YES:completed", ""],
            [MACHINE_TYPE_CODE, "LMD_VALVE_MAINT", "阀维修前需要安全隔离", "", "true", "LMD_SAFE_YES:completed", ""],
            [MACHINE_TYPE_CODE, "LMD_ACCEPTANCE", "验收前需要维修完成", "", "true", "LMD_REPAIR_DONE:completed", ""],
        ],
    )

    _append_rows(
        workbook["rules"],
        [
            [
                "OP_LMD_MAKE_SAFE",
                MACHINE_TYPE_CODE,
                "执行安全隔离",
                5,
                "Provide the shared safety fact once.",
                "true",
                "false",
                "",
                "safe:set:yes",
                "TECHNICIAN:1:true",
                "LMD_ISOLATE",
            ],
            [
                "OP_LMD_DIAGNOSE_PUMP",
                MACHINE_TYPE_CODE,
                "诊断泵故障",
                4,
                "Pump maintenance step 1.",
                "true",
                "true",
                "",
                "pump_diagnosed:set:yes",
                "TECHNICIAN:1:true",
                "LMD_PUMP_DIAGNOSE",
            ],
            [
                "OP_LMD_FIX_PUMP",
                MACHINE_TYPE_CODE,
                "修复泵",
                7,
                "Pump maintenance step 2.",
                "true",
                "true",
                "pump_diagnosed:eq:yes",
                "pump_fixed:set:yes",
                "TECHNICIAN:1:true",
                "LMD_PUMP_REPAIR",
            ],
            [
                "OP_LMD_DIAGNOSE_VALVE",
                MACHINE_TYPE_CODE,
                "诊断阀故障",
                3,
                "Valve maintenance step 1.",
                "true",
                "true",
                "",
                "valve_diagnosed:set:yes",
                "TECHNICIAN:1:true",
                "LMD_VALVE_DIAGNOSE",
            ],
            [
                "OP_LMD_FIX_VALVE",
                MACHINE_TYPE_CODE,
                "修复阀",
                8,
                "Valve maintenance step 2.",
                "true",
                "true",
                "valve_diagnosed:eq:yes",
                "valve_fixed:set:yes",
                "TECHNICIAN:1:true",
                "LMD_VALVE_REPAIR",
            ],
            [
                "OP_LMD_FINAL_TEST",
                MACHINE_TYPE_CODE,
                "执行联合验收",
                4,
                "Final acceptance after repair completion.",
                "true",
                "false",
                "",
                "final_test:set:passed",
                "TECHNICIAN:1:true",
                "LMD_FINAL_TEST",
            ],
        ],
    )

    _append_rows(
        workbook["maintenance_intents"],
        [
            [
                MACHINE_TYPE_CODE,
                "pump_fault",
                "泵维修维护",
                "LMD_PUMP_MAINT",
                "Restore pump health through the minimal pump maintenance capability set.",
                "LMD_PUMP_FIXED_YES",
                "LMD_PREP;LMD_PUMP_MAINT",
                "pump_diagnosed:eq:no;pump_fixed:eq:no",
                "",
                "true",
                "",
            ],
            [
                MACHINE_TYPE_CODE,
                "valve_fault",
                "阀维修维护",
                "LMD_VALVE_MAINT",
                "Restore valve health through the minimal valve maintenance capability set.",
                "LMD_VALVE_FIXED_YES",
                "LMD_PREP;LMD_VALVE_MAINT",
                "valve_diagnosed:eq:no;valve_fixed:eq:no",
                "",
                "true",
                "",
            ],
        ],
    )

    start_features = (
        "safe:no;"
        "pump_diagnosed:no;pump_fixed:no;"
        "valve_diagnosed:no;valve_fixed:no;"
        "final_test:not_run"
    )
    target_features = (
        "safe:yes;"
        "pump_diagnosed:yes;pump_fixed:yes;"
        "valve_diagnosed:yes;valve_fixed:yes;"
        "final_test:passed"
    )
    _append_rows(
        workbook["states"],
        [
            [MACHINE_CODE, "LMD_START", "current", "分层维护起始状态", start_features],
            [MACHINE_CODE, "LMD_TARGET", "target", "分层维护目标状态", target_features],
        ],
    )
    _append_rows(
        workbook["solve_cases"],
        [[
            "LMD_FULL_LAYERED_FLOW",
            MACHINE_CODE,
            "LMD_START",
            "LMD_TARGET",
            "minimize_makespan",
            "",
            "",
            6,
            20,
        ]],
    )
    _append_rows(
        workbook["layered_health_checks"],
        [[
            MACHINE_TYPE_CODE,
            "LMD_FULL_HEALTH",
            "Full layered Provider/Consumer health",
            "LMD_OPERATION_READY",
            "LMD_SYSTEM",
            "false",
            "Validate root state aggregation, scope guards, and all providers.",
        ]],
    )
    _append_rows(
        workbook["notes"],
        [["Fixture validates layered state aggregation, cross-level scope guards, maintenance merge, and continuity diagnostics."]],
    )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def _load_ids(session: Any, import_result: dict[str, Any]) -> dict[str, Any]:
    machine = (
        await session.execute(select(Machine).where(Machine.code == MACHINE_CODE).limit(1))
    ).scalar_one()
    state_case = import_result["solve_cases"][0]
    current_state_id = state_case["current_state_id"]

    state_nodes = {
        row.code: row.id
        for row in (
            await session.execute(
                select(StateNode).where(StateNode.machine_type_id == machine.machine_type_id)
            )
        ).scalars()
    }
    activity_nodes = {
        row.code: row.id
        for row in (
            await session.execute(
                select(ActivityNode).where(ActivityNode.machine_type_id == machine.machine_type_id)
            )
        ).scalars()
    }
    templates = {
        row.issue_type: row.id
        for row in (
            await session.execute(
                select(MaintenanceIntentTemplate).where(
                    MaintenanceIntentTemplate.machine_type_id == machine.machine_type_id,
                    MaintenanceIntentTemplate.issue_type.in_(["pump_fault", "valve_fault"]),
                )
            )
        ).scalars()
    }
    return {
        "machine_id": machine.id,
        "machine_type_id": machine.machine_type_id,
        "current_state_id": current_state_id,
        "state_nodes": state_nodes,
        "activity_nodes": activity_nodes,
        "maintenance_templates": templates,
    }


def _task_codes(result: dict[str, Any]) -> list[str]:
    return [task["op_rule_code"] for task in result.get("schedule", {}).get("tasks", [])]


def _state_features(result: dict[str, Any]) -> dict[str, Any]:
    return result.get("layered", {}).get("state_replay", {}).get("final_state", {})


def _selection_by_code(result: dict[str, Any]) -> dict[str, Any]:
    rows = result.get("layered", {}).get("activity_selection", [])
    return {row.get("op_rule_code"): row for row in rows if row.get("op_rule_code")}


async def verify_import() -> dict[str, Any]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook_bytes = build_workbook()
    XLSX_PATH.write_bytes(workbook_bytes)

    parsed = parse_scenario_workbook(workbook_bytes)
    async with AsyncSessionLocal() as session:
        validation = await validate_scenario_workbook(parsed, session)
        if validation.get("errors"):
            await session.rollback()
            return {
                "status": "validation_failed",
                "xlsx_path": str(XLSX_PATH),
                "validation": validation,
            }

        import_result = await import_scenario_workbook(parsed, session)
        await session.commit()

    async with AsyncSessionLocal() as session:
        ids = await _load_ids(session, import_result)

        layered = await solve_layered(
            LayeredSolveRequest(
                machine_id=ids["machine_id"],
                current_state_id=ids["current_state_id"],
                target_state_node_ids=[ids["state_nodes"]["LMD_OPERATION_READY"]],
                activity_scope_node_ids=[ids["activity_nodes"]["LMD_SYSTEM"]],
                objectives=[
                    {"type": "minimize_makespan", "weight": 1.0},
                    {"type": "minimize_activity_group_span", "weight": 1.0},
                    {"type": "minimize_activity_group_gaps", "weight": 1.0},
                    {"type": "minimize_activity_group_interruptions", "weight": 1.0},
                ],
            ),
            session,
        )

    async with AsyncSessionLocal() as session:
        ids = await _load_ids(session, import_result)
        maintenance = await solve_maintenance(
            MaintenanceSolveRequest(
                machine_id=ids["machine_id"],
                current_state_id=ids["current_state_id"],
                intent_template_ids=[
                    ids["maintenance_templates"]["pump_fault"],
                    ids["maintenance_templates"]["valve_fault"],
                ],
                objectives=[
                    {"type": "minimize_makespan", "weight": 1.0},
                    {"type": "minimize_activity_group_span", "weight": 1.0},
                    {"type": "minimize_activity_group_gaps", "weight": 1.0},
                    {"type": "minimize_activity_group_interruptions", "weight": 1.0},
                ],
            ),
            session,
        )

    async with AsyncSessionLocal() as session:
        ids = await _load_ids(session, import_result)
        maintenance_already_safe = await solve_maintenance(
            MaintenanceSolveRequest(
                machine_id=ids["machine_id"],
                current_state_id=ids["current_state_id"],
                intent_template_ids=[
                    ids["maintenance_templates"]["pump_fault"],
                    ids["maintenance_templates"]["valve_fault"],
                ],
                extra_observed_facts=[
                    {"feature_key": "safe", "operator": "eq", "value": "yes"},
                ],
                objectives=[{"type": "minimize_makespan", "weight": 1.0}],
            ),
            session,
        )

    layered_tasks = _task_codes(layered)
    maintenance_tasks = _task_codes(maintenance)
    already_safe_tasks = _task_codes(maintenance_already_safe)
    selection = _selection_by_code(maintenance)
    safe_selection = selection.get("OP_LMD_MAKE_SAFE") or {}

    assertions = {
        "layered_solve_done": layered.get("status") == "done",
        "layered_root_expands_to_6_goal_facts": len(layered.get("layered", {}).get("goal_facts", [])) == 6,
        "layered_runs_full_sequence_once": layered_tasks == [
            "OP_LMD_MAKE_SAFE",
            "OP_LMD_DIAGNOSE_VALVE",
            "OP_LMD_DIAGNOSE_PUMP",
            "OP_LMD_FIX_VALVE",
            "OP_LMD_FIX_PUMP",
            "OP_LMD_FINAL_TEST",
        ] or sorted(layered_tasks) == sorted([
            "OP_LMD_MAKE_SAFE",
            "OP_LMD_DIAGNOSE_PUMP",
            "OP_LMD_FIX_PUMP",
            "OP_LMD_DIAGNOSE_VALVE",
            "OP_LMD_FIX_VALVE",
            "OP_LMD_FINAL_TEST",
        ]),
        "layered_final_state_complete": all(
            _state_features(layered).get(key) == value
            for key, value in {
                "safe": "yes",
                "pump_diagnosed": "yes",
                "pump_fixed": "yes",
                "valve_diagnosed": "yes",
                "valve_fixed": "yes",
                "final_test": "passed",
            }.items()
        ),
        "maintenance_solve_done": maintenance.get("status") == "done",
        "maintenance_merges_two_intents": maintenance.get("maintenance", {}).get("merged_intent_count") == 2,
        "maintenance_shared_safety_once": maintenance_tasks.count("OP_LMD_MAKE_SAFE") == 1,
        "maintenance_shared_provider_explained": bool(safe_selection.get("is_shared_provider")),
        "maintenance_continuity_diagnostics_present": bool(
            maintenance.get("diagnostics", {}).get("schedule", {}).get("activity_group_continuity")
        ),
        "already_safe_solve_done": maintenance_already_safe.get("status") == "done",
        "already_safe_skips_safety_step": (
            maintenance_already_safe.get("status") == "done"
            and "OP_LMD_MAKE_SAFE" not in already_safe_tasks
        ),
    }

    summary = {
        "status": "ok" if all(assertions.values()) else "failed",
        "xlsx_path": str(XLSX_PATH),
        "summary_path": str(SUMMARY_PATH),
        "validation": validation,
        "import_result": import_result,
        "ids": {
            "machine_id": ids["machine_id"],
            "machine_type_id": ids["machine_type_id"],
            "current_state_id": ids["current_state_id"],
            "maintenance_templates": ids["maintenance_templates"],
        },
        "assertions": assertions,
        "layered": {
            "status": layered.get("status"),
            "solve_request_id": layered.get("solve_request_id"),
            "makespan": layered.get("schedule", {}).get("makespan"),
            "task_codes": layered_tasks,
            "goal_fact_count": len(layered.get("layered", {}).get("goal_facts", [])),
            "final_state": _state_features(layered),
            "state_tree": layered.get("layered", {}).get("state_tree"),
        },
        "maintenance": {
            "status": maintenance.get("status"),
            "solve_request_id": maintenance.get("solve_request_id"),
            "makespan": maintenance.get("schedule", {}).get("makespan"),
            "task_codes": maintenance_tasks,
            "merged_intent_count": maintenance.get("maintenance", {}).get("merged_intent_count"),
            "shared_safety_selection": safe_selection,
            "continuity": maintenance.get("diagnostics", {}).get("schedule", {}).get("activity_group_continuity"),
        },
        "maintenance_already_safe": {
            "status": maintenance_already_safe.get("status"),
            "solve_request_id": maintenance_already_safe.get("solve_request_id"),
            "makespan": maintenance_already_safe.get("schedule", {}).get("makespan"),
            "task_codes": already_safe_tasks,
            "current_state_overrides": maintenance_already_safe.get("layered", {}).get("current_state_overrides"),
        },
    }
    SUMMARY_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main() -> None:
    summary = asyncio.run(verify_import())
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if summary["status"] != "ok":
        raise SystemExit(1)


if __name__ == "__main__":
    main()
